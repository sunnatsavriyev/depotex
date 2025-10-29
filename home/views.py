from rest_framework import viewsets, status, filters, mixins, generics
from rest_framework.permissions import IsAuthenticated, AllowAny,SAFE_METHODS
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.viewsets import ViewSet
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from .models import (
    TamirTuri, ElektroDepo, EhtiyotQismlari,
    HarakatTarkibi, TexnikKorik, CustomUser, Nosozliklar, NosozlikEhtiyotQism, TexnikKorikStep,KunlikYurish,
    Vagon,EhtiyotQismHistory, TexnikKorikEhtiyotQism,NosozlikTuri,TexnikKorikJadval,Notification,Marshrut,YilOy
)
from .serializers import (
    TamirTuriSerializer, ElektroDepoSerializer,TexnikKorikJadvalSerializer,
    EhtiyotQismlariSerializer, HarakatTarkibiSerializer,
    TexnikKorikSerializer, UserSerializer, NosozliklarSerializer, TexnikKorikStepSerializer, NosozlikStepSerializer,MarshrutSerializer,
    NosozlikStep,KunlikYurishSerializer,VagonSerializer,NosozlikTuriSerializer,NotificationSerializer,YilOySerializer,
    HarakatTarkibiActiveSerializer, EhtiyotQismWithMiqdorSerializer,EhtiyotQismHistorySerializer, TarkibFullDetailSerializer,TexnikKorikDetailForStepSerializer,NosozlikDetailForStepSerializer)
from django.utils import timezone
from django.db.models import Sum, F
from django.contrib.auth import authenticate
from .pagination import CustomPagination
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.exceptions import PermissionDenied
from rest_framework.decorators import api_view, permission_classes, action
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape, inch
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, HRFlowable, Image,PageBreak, Flowable
import io
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.graphics.shapes import Drawing, Circle, Polygon, Rect, String
from reportlab.graphics import renderPDF
import requests
import django_filters
from rest_framework.exceptions import ValidationError
from reportlab.platypus import Paragraph, Spacer, HRFlowable, Image, Table, TableStyle
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from django.db.models import Count
from django.utils.timezone import now
from datetime import timedelta, datetime, date
from rest_framework.views import APIView
from drf_spectacular.utils import extend_schema, OpenApiParameter,OpenApiExample
from .permissions import IsMonitoringReadOnly, IsTexnik, IsJadvalchi
import json
from reportlab.lib.units import cm
from drf_yasg.utils import swagger_auto_schema
from drf_spectacular.utils import extend_schema, OpenApiParameter, OpenApiTypes
from drf_yasg import openapi
from io import BytesIO
import qrcode
from collections import defaultdict
import pandas as pd
import calendar
from reportlab.lib.pagesizes import A3, landscape
from django_filters.rest_framework import DjangoFilterBackend, FilterSet, DateFromToRangeFilter, NumberFilter
from reportlab.lib.enums import TA_RIGHT
class UserViewSet(viewsets.ModelViewSet):
    queryset = CustomUser.objects.all()
    serializer_class = UserSerializer
    permission_classes = [IsAuthenticated]



@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_me(request):
    user = request.user
    
    # Headerdan tokenni olish
    auth_header = request.headers.get("Authorization", None)
    token = None
    if auth_header and auth_header.startswith("Bearer "):
        token = auth_header.split(" ")[1]
        
        
    today = date.today()
    todays_checks = TexnikKorikJadval.objects.filter(sana=today)

    for check in todays_checks:
        depo = getattr(check.tarkib, "depo", None)
        if depo and user in depo.users.filter(role="texnik"):
            message = (
                f"Bugun {check.tarkib.tarkib_raqami} tarkibi uchun "
                f"'{check.tamir_turi.tamir_nomi}' texnik ko‘rik rejalashtirilgan."
            )

            Notification.objects.get_or_create(
                user=user,
                title="Bugungi texnik ko‘rik eslatmasi",
                message=message,
                defaults={"is_read": False},
            )

    return Response({
        "id": user.id,
        "username": user.username,
        "role": user.role,
        "token": token,  
    })


class BaseViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated,IsMonitoringReadOnly,IsTexnik]
    require_login_fields = False
    filter_backends = [filters.SearchFilter, filters.OrderingFilter, DjangoFilterBackend]
    search_fields = []  
    ordering_fields = "__all__"
    
    def get_queryset(self):
        if self.queryset is not None:
            return self.queryset
        return self.serializer_class.Meta.model.objects.none()

    def get_serializer_context(self):
        context = super().get_serializer_context()
        context["request"] = self.request
        return context

    def create(self, request, *args, **kwargs):
        # Faqat require_login_fields = True bo'lsa tekshiramiz
        if getattr(self, "require_login_fields", False):
            username = request.data.get("username")
            password = request.data.get("password")

            if not username or not password:
                return Response(
                    {"detail": "Username va parol talab qilinadi."},
                    status=status.HTTP_400_BAD_REQUEST
                )

            user = authenticate(username=username, password=password)
            if not user:
                return Response(
                    {"detail": "Username yoki parol xato."},
                    status=status.HTTP_403_FORBIDDEN
                )

            context = {"request": request}
        else:
            context = {"request": request}  

        serializer = self.get_serializer(data=request.data, context=context)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
        return Response(serializer.data, status=status.HTTP_201_CREATED)
    

    def generate_pdf_detail(self, filename, title, data_list):
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, title=title)

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle('title_style', parent=styles['Title'], textColor=colors.darkblue, alignment=1)
        header_style = ParagraphStyle('header_style', parent=styles['Heading2'], textColor=colors.white, backColor=colors.darkblue, alignment=1)
        field_style = ParagraphStyle('field_style', parent=styles['Normal'], textColor=colors.black, fontSize=9, leading=11)

        elements = []
        elements.append(Paragraph(title, title_style))
        elements.append(Spacer(1, 20))

        for idx, item in enumerate(data_list, start=1):
            elements.append(Paragraph(f"Obyekt #{idx}", header_style))
            elements.append(Spacer(1, 10))

            # Asosiy jadval
            table_data = [[Paragraph("<b>Maydon</b>", field_style), Paragraph("<b>Qiymat</b>", field_style)]]
            for key, value in item.items():
                if key in ["image", "steps", "vagonlar", "ehtiyot_qismlar_detail"]:
                    continue
                table_data.append([Paragraph(str(key), field_style), Paragraph(str(value), field_style)])

            # Ehtiyot qismlarni birlashtirib yozish
            if "ehtiyot_qismlar_detail" in item and isinstance(item["ehtiyot_qismlar_detail"], list):
                qismlar_list = []
                for part in item["ehtiyot_qismlar_detail"]:
                    nomi = part.get("ehtiyot_qism_nomi", "")
                    miqdor = part.get("ishlatilgan_miqdor", 0)
                    birligi = part.get("birligi", "")
                    qismlar_list.append(f"{nomi}: {miqdor} {birligi}")
                table_data.append([Paragraph("Ehtiyot qismlar", field_style), Paragraph(", ".join(qismlar_list), field_style)])

            table = Table(table_data, colWidths=[150, 350])
            table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ]))
            elements.append(table)
            elements.append(Spacer(1, 15))

            # Steps jadvali (alohida)
            if "steps" in item and isinstance(item["steps"], dict):
                steps = item["steps"].get("results", [])
                if steps:
                    elements.append(Paragraph("🔹 Ko‘rik bosqichlari", header_style))
                    elements.append(Spacer(1, 8))

                    step_table_data = [
                        [Paragraph("<b>ID</b>", field_style),
                        Paragraph("<b>Kamchiliklar</b>", field_style),
                        Paragraph("<b>Kim tomonidan</b>", field_style)]
                    ]

                    for step in steps:
                        step_table_data.append([
                            Paragraph(str(step.get("id")), field_style),
                            Paragraph(str(step.get("kamchiliklar_haqida")), field_style),
                            Paragraph(str(step.get("created_by")), field_style),
                        ])

                    step_table = Table(step_table_data, colWidths=[40, 200, 150])
                    step_table.setStyle(TableStyle([
                        ("BACKGROUND", (0, 0), (-1, 0), colors.lightblue),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
                        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                        ("FONTSIZE", (0, 0), (-1, -1), 8),
                        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                    ]))
                    elements.append(step_table)
                    elements.append(Spacer(1, 20))

            elements.append(HRFlowable(width="100%", color=colors.darkgrey, thickness=0.7))
            elements.append(Spacer(1, 20))

        doc.build(elements)
        buffer.seek(0)

        response = HttpResponse(buffer, content_type="application/pdf")
        response["Content-Disposition"] = f'attachment; filename="{filename}.pdf"'
        return response



    # 🔹 PDF eksport
    @action(detail=False, methods=["get"], url_path="export-pdf")
    def export_pdf(self, request):
        queryset = self.get_queryset()
        serializer = self.get_serializer(queryset, many=True)
        data_list = serializer.data
        return self.generate_pdf_detail(self.basename, f"{self.basename.capitalize()} ro'yxati", data_list)



    excel_headers = None        # optional: ["ID", "Name", ...]
    excel_filename = None       # optional: "my_export.xlsx"
    exclude_excel_fields = ["image", "vagonlar", "steps"]   # keraksiz ustunlar

    @action(detail=False, methods=["get"], url_path="export-excel")
    def export_excel(self, request):
        queryset = self.filter_queryset(self.get_queryset())
        serializer = self.get_serializer(queryset, many=True)

        wb = Workbook()
        ws = wb.active
        ws.title = (getattr(self, "basename", None) or self.__class__.__name__).capitalize()

        # Headerlar
        headers = self.get_excel_headers(serializer)
        if headers:
            ws.append(headers)

            # Header style
            header_fill = PatternFill("solid", fgColor="4F81BD")
            header_font = Font(bold=True, color="FFFFFF")
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )

        # Data yozish
        for row_data in self.format_excel_data(serializer.data, headers=headers):
            ws.append(row_data)

        # Border & alignment
        if ws.max_row >= 2:
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
                for cell in row:
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical="center", wrap_text=True)

        # Column auto width
        for col in ws.columns:
            try:
                max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
                ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2
            except Exception:
                pass

        # Fayl nomi
        filename = self.excel_filename or f"{(getattr(self, 'basename', None) or self.__class__.__name__).lower()}.xlsx"
        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response

    def get_excel_headers(self, serializer):
        """
        Agar child viewset excel_headers bersa — uni ishlatamiz.
        Aks holda serializer.data[0].keys() dan olamiz va exclude qilamiz.
        """
        headers = []
        if getattr(self, "excel_headers", None):
            headers = list(self.excel_headers)
        elif getattr(serializer, "data", None):
            if len(serializer.data) > 0 and isinstance(serializer.data[0], dict):
                headers = list(serializer.data[0].keys())

        # exclude qilingan ustunlarni olib tashlaymiz
        exclude = getattr(self, "exclude_excel_fields", [])
        headers = [h for h in headers if h not in exclude]
        return headers

    def format_excel_data(self, data_list, headers=None):
        """
        Serializer.data ichidan faqat kerakli headerlarga mos qiymatlarni olib beradi.
        """
        rows = []
        headers = headers or []
        if not data_list:
            return rows

        for obj in data_list:
            row = []
            for h in headers:
                value = obj.get(h, "")
                if isinstance(value, (list, dict)):
                    try:
                        value = json.dumps(value, ensure_ascii=False)
                    except Exception:
                        value = str(value)
                row.append(value if value is not None else "")
            rows.append(row)
        return rows


class TamirTuriViewSet(BaseViewSet):
    queryset = TamirTuri.objects.all()
    serializer_class = TamirTuriSerializer
    basename = "Tamir Turi"
    permission_classes = [IsAuthenticated,IsTexnik]
    require_login_fields = False
    filter_backends = [filters.SearchFilter, filters.OrderingFilter, DjangoFilterBackend]
    search_fields = ['tamir_nomi',"tarkib_turi", 'tamirlash_davri', 'tamirlanish_vaqti']   
    ordering_fields = ['tamirlanish_vaqti', 'id']
    pagination_class = CustomPagination
    


class ElektroDepoViewSet(BaseViewSet):
    queryset = ElektroDepo.objects.all()
    serializer_class = ElektroDepoSerializer
    basename = "Elektro Depo"
    permission_classes = [IsAuthenticated, IsTexnik]
    require_login_fields = False
    filter_backends = [filters.SearchFilter, filters.OrderingFilter, DjangoFilterBackend]
    search_fields = ['depo_nomi', 'qisqacha_nomi', 'joylashuvi']   
    ordering_fields = ['qisqacha_nomi', 'id']
    def get_permissions(self):
        # READ (detail view) uchun public; boshqa actionlar uchun normal role-based permissionlar qoʻyiladi
        if self.action in ['retrieve', 'list']:
            return [AllowAny()]
        return [IsAuthenticated()] 
    

class EhtiyotQismlariViewSet(viewsets.ModelViewSet):
    queryset = EhtiyotQismlari.objects.all().order_by('-id')
    serializer_class = EhtiyotQismlariSerializer
    permission_classes = [IsAuthenticated, IsTexnik]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['ehtiyotqism_nomi', 'nomenklatura_raqami']
    ordering_fields = ['id', 'nomenklatura_raqami']
    pagination_class = CustomPagination
    
    
class EhtiyotQismMiqdorListAPIView(APIView):
        permission_classes = [IsAuthenticated]

        def get(self, request, ehtiyotqism_pk):
            try:
                ehtiyot_qism = EhtiyotQismlari.objects.get(pk=ehtiyotqism_pk)
            except EhtiyotQismlari.DoesNotExist:
                return Response({"error": "Ehtiyot qism topilmadi"}, status=status.HTTP_404_NOT_FOUND)

            history = EhtiyotQismHistory.objects.filter(
                ehtiyot_qism=ehtiyot_qism
            ).order_by('-created_at')

            jami_miqdor = history.aggregate(total=Sum('miqdor'))['total'] or 0
            serializer = EhtiyotQismHistorySerializer(history, many=True)

            return Response({
                "id": ehtiyot_qism.id,
                "ehtiyotqism_nomi": ehtiyot_qism.ehtiyotqism_nomi,
                "birligi": ehtiyot_qism.birligi,
                "depo": ehtiyot_qism.depo.qisqacha_nomi,
                "jami_miqdor": jami_miqdor,
                "history": serializer.data
            })

class EhtiyotQismMiqdorCreateAPIView(generics.CreateAPIView):
    serializer_class = EhtiyotQismHistorySerializer
    permission_classes = [IsAuthenticated, IsTexnik]

    def perform_create(self, serializer):
        ehtiyotqism_pk = self.kwargs.get("ehtiyotqism_pk")
        ehtiyot_qism = EhtiyotQismlari.objects.get(pk=ehtiyotqism_pk)

        serializer.save(
            ehtiyot_qism=ehtiyot_qism,
            created_by=self.request.user
        )

    def create(self, request, *args, **kwargs):
        response = super().create(request, *args, **kwargs)

        ehtiyotqism_pk = self.kwargs.get("ehtiyotqism_pk")
        ehtiyot_qism = EhtiyotQismlari.objects.get(pk=ehtiyotqism_pk)

        history = EhtiyotQismHistory.objects.filter(
            ehtiyot_qism=ehtiyot_qism
        ).order_by("-created_at")

        jami_miqdor = history.aggregate(total=Sum("miqdor"))['total'] or 0

        response.data = {
            "status": "Miqdor qo'shildi",
            "id": ehtiyot_qism.id,
            "jami_miqdor": jami_miqdor,
            "history": EhtiyotQismHistorySerializer(history, many=True).data,
        }
        return response
        

class HarakatTarkibiViewSet(BaseViewSet):
    queryset = (
    HarakatTarkibi.objects.filter(is_active=True)
    .annotate(total_kilometr=Sum("kunlik_yurishlar__kilometr"))
    .order_by("-id")
    )
    serializer_class = HarakatTarkibiSerializer
    basename = "Harakat Tarkibi"
    permission_classes = [IsAuthenticated, IsTexnik]
    require_login_fields = False
    def get_queryset(self):
        user = self.request.user

        # superuser hammasini ko‘radi
        if user.is_superuser:
            return HarakatTarkibi.objects.all().order_by("-id")

        # texnik faqat o‘z deposidagi tarkiblarni ko‘radi
        if user.role == "texnik" and user.depo:
            return HarakatTarkibi.objects.all().order_by("-id")

        # monitoring hammasini faqat o‘qiy oladi
        if user.role == "monitoring":
            return HarakatTarkibi.objects.all().order_by("-id")

        return HarakatTarkibi.objects.none()
    pagination_class = CustomPagination
    filter_backends = [filters.SearchFilter, filters.OrderingFilter, DjangoFilterBackend]
    search_fields = ['guruhi','tarkib_raqami','turi','ishga_tushgan_vaqti','eksplutatsiya_vaqti','holati']   
    ordering_fields = ['ishga_tushgan_vaqti', 'id']
    



class HarakatTarkibiGetViewSet(BaseViewSet):
    queryset = (
    HarakatTarkibi.objects.filter(is_active=True)
    .annotate(total_kilometr=Sum("kunlik_yurishlar__kilometr"))
    .order_by("-id")
    )
    serializer_class = HarakatTarkibiSerializer
    basename = "Harakat Tarkibi"
    permission_classes = [IsAuthenticated, IsTexnik]
    require_login_fields = False
    pagination_class = CustomPagination
    filter_backends = [filters.SearchFilter, filters.OrderingFilter, DjangoFilterBackend]
    search_fields = ['guruhi', 'tarkib_raqami', 'turi', 'ishga_tushgan_vaqti', 'eksplutatsiya_vaqti','holati']
    ordering_fields = ['ishga_tushgan_vaqti', 'id']
    filterset_fields = ['depo']


    @action(detail=False, methods=["get"], url_path="by-depo")
    def by_depo(self, request):
        depo_id = request.query_params.get("depo_id")
        if not depo_id:
            return Response({"error": "depo_id query param kerak"}, status=400)

        queryset = (
            self.get_queryset()
            .filter(depo_id=depo_id, is_active=True)
        )

        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)

        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)
    
    
    @action(detail=True, methods=["get"], url_path="versions")
    def versions(self, request, pk=None):
        obj = self.get_object()
        versions = HarakatTarkibi.objects.filter(
            tarkib_raqami=obj.tarkib_raqami
        ).order_by("-id")

        serializer = self.get_serializer(versions, many=True)
        return Response(serializer.data)


    def perform_create(self, serializer):
        # Agar tarkib_raqami bo‘yicha eski versiya bo‘lsa → uni deactivate qilamiz
        tarkib_raqami = serializer.validated_data.get("tarkib_raqami")
        eski_versiya = (
            HarakatTarkibi.objects.filter(tarkib_raqami=tarkib_raqami, is_active=True)
            .order_by("-id")
            .first()
        )

        if eski_versiya:
            eski_versiya.is_active = False
            eski_versiya.save()

            # yangi obyekt eski versiyaga bog‘lanadi
            serializer.save(
                created_by=self.request.user,
                previous_version=eski_versiya,
                is_active=True,
            )
        else:
            # birinchi versiya bo‘lsa
            serializer.save(created_by=self.request.user, is_active=True)
            
    @action(detail=True, methods=["get"], url_path="vagonlar")
    def vagonlar(self, request, pk=None):
        obj = self.get_object()
        vagonlar = obj.vagon_set.all()  # yoki obj.vagonlar.all() agar related_name="vagonlar"
        serializer = VagonSerializer(vagonlar, many=True)
        return Response(serializer.data)



class HarakatTarkibiActiveViewSet(BaseViewSet):
    queryset = (
        HarakatTarkibi.objects.filter(is_active=True)
        .annotate(total_kilometr=Sum("kunlik_yurishlar__kilometr"))
        .order_by("-id")
    )
    serializer_class = HarakatTarkibiActiveSerializer
    basename = "harakat-tarkibi-active"  #  Bo‘sh joysiz
    permission_classes = [IsAuthenticated, IsTexnik]
    require_login_fields = False
    pagination_class = CustomPagination
    filter_backends = [filters.SearchFilter, filters.OrderingFilter, DjangoFilterBackend]
    search_fields = ['guruhi', 'tarkib_raqami', 'turi', 'ishga_tushgan_vaqti', 'eksplutatsiya_vaqti', 'holati']
    ordering_fields = ['ishga_tushgan_vaqti', 'id']
    filterset_fields = ['depo']
    
    
    def get_queryset(self):
        user = self.request.user
        
        # SuperUser uchun barcha depolardagi tarkiblarni ko'rsatish
        if user.is_superuser:
            queryset = HarakatTarkibi.objects.filter(is_active=True)
        # Oddiy user uchun faqat o'z deposidagi tarkiblarni ko'rsatish
        else:
            user_depo = user.depo
            if user_depo:
                queryset = HarakatTarkibi.objects.filter(is_active=True, depo=user_depo)
            else:
                queryset = HarakatTarkibi.objects.none()
        
        return queryset



class HarakatTarkibiHolatStatistikaViewSet(viewsets.GenericViewSet):
    permission_classes = [IsAuthenticated, IsTexnik]

    # --- Asosiy statistikani chiqarish ---
    def list(self, request):
        queryset = HarakatTarkibi.objects.filter(is_active=True)

        nosozlikda = queryset.filter(holati="Nosozlikda")
        texnik_korikda = queryset.filter(holati="Texnik_korikda")

        nosozlik_serializer = HarakatTarkibiActiveSerializer(nosozlikda, many=True)
        texnik_serializer = HarakatTarkibiActiveSerializer(texnik_korikda, many=True)

        return Response({
            "nosozlikda_soni": nosozlikda.count(),
            "nosozlikda_tarkiblar": nosozlik_serializer.data,
            "texnik_korikda_soni": texnik_korikda.count(),
            "texnik_korikda_tarkiblar": texnik_serializer.data,
        })

    # --- PDF EXPORT ---
    @extend_schema(
    parameters=[
        OpenApiParameter(
            name="type",
            description="Qaysi turdagi PDF kerak?",
            required=True,
            type=str,
            examples=[
                OpenApiExample(
                    name="Nosozlik",
                    summary="Nosozlikda bo‘lgan tarkiblar",
                    value="nosozlikda"
                ),
                OpenApiExample(
                    name="Texnik",
                    summary="Texnik ko‘rikda bo‘lgan tarkiblar",
                    value="texnik"
                ),
            ],
        ),
    ],
    responses={200: {"description": "PDF fayl qaytaradi"}}
    )
    @action(detail=False, methods=["get"], url_path="export-pdf")
    def export_pdf(self, request):
        pdf_type = request.query_params.get("type", "").lower().strip("/")

        if pdf_type not in ["nosozlikda", "texnik"]:
            return Response({
                "detail": "Iltimos, ?type=nosozlikda yoki ?type=texnik deb kiriting."
            }, status=400)

        # --- Filter va sozlamalar ---
        if pdf_type == "nosozlikda":
            queryset = HarakatTarkibi.objects.filter(is_active=True, holati="Nosozlikda")
            title_text = "Nosozlikda bo‘lgan tarkiblar ro‘yxati"
            title_color = "red"
            filename_prefix = "Nosozlikda_Tarkiblar"
        else:
            queryset = HarakatTarkibi.objects.filter(is_active=True, holati="Texnik_korikda")
            title_text = "Texnik ko‘rikda bo‘lgan tarkiblar ro‘yxati"
            title_color = "blue"
            filename_prefix = "TexnikKorik_Tarkiblar"

        # --- PDF yaratish ---
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=1*cm, bottomMargin=1*cm)
        elements = []

        title = Paragraph(
            f"<b><font size=14 color='{title_color}'>{title_text}</font></b>",
            ParagraphStyle('title', alignment=1)
        )
        elements += [title, Spacer(1, 10)]

        data = [["#", "Tarkib raqami", "Turi", "Masofa", "Holati"]]
        for i, obj in enumerate(queryset, 1):
            data.append([
                i,
                str(obj.tarkib_raqami or "-"),
                str(obj.turi or "-"),
                str(getattr(obj, 'masofa', "-")),
                str(obj.holati or "-"),
            ])

        table = Table(data, hAlign='CENTER', colWidths=[1*cm, 4*cm, 4*cm, 3*cm, 5*cm])
        table.setStyle(TableStyle([
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('BACKGROUND', (0, 0), (-1, 0), colors.whitesmoke),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.darkblue),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ]))
        elements.append(table)

        doc.build(elements)
        pdf = buffer.getvalue()
        buffer.close()

        filename = f"{filename_prefix}_{datetime.now().strftime('%Y%m%d')}.pdf"
        response = HttpResponse(pdf, content_type="application/pdf")
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response


    # --- EXCEL EXPORT ---
    
    @action(detail=False, methods=["get"], url_path="export-excel")
    def export_excel(self, request):
        queryset = HarakatTarkibi.objects.filter(is_active=True)
        nosozlikda = queryset.filter(holati="Nosozlikda")
        texnik_korikda = queryset.filter(holati="Texnik_korikda")  # ✅ diqqat: modeldagi aniq qiymatni yozing

        # Excel fayl yaratish
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Nosozlikda"
        ws2 = wb.create_sheet("Texnik ko‘rikda")

        headers = ["#", "Tarkib raqami", "Turi", "Masofa", "Holati"]

        # --- Nosozlikda sahifasi ---
        ws1.append(headers)
        for i, obj in enumerate(nosozlikda, 1):
            ws1.append([
                i,
                str(obj.tarkib_raqami or "-"),
                str(obj.turi or "-"),
                str(getattr(obj, 'masofa', "-")),
                str(obj.holati or "-"),
            ])

        # --- Texnik ko‘rikda sahifasi ---
        ws2.append(headers)
        for i, obj in enumerate(texnik_korikda, 1):
            ws2.append([
                i,
                str(obj.tarkib_raqami or "-"),
                str(obj.turi or "-"),
                str(getattr(obj, 'masofa', "-")),
                str(obj.holati or "-"),
            ])

        # Ustun kengliklarini moslashtirish
        for sheet in [ws1, ws2]:
            for col in range(1, len(headers) + 1):
                sheet.column_dimensions[get_column_letter(col)].width = 20

        # Excel faylni xotirada saqlash
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Javobga yuborish
        response = HttpResponse(
            output,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        filename = f"HarakatTarkibi_Holatlar_{datetime.now().strftime('%Y%m%d')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response




class KunlikYurishViewSet(BaseViewSet):
    serializer_class = KunlikYurishSerializer
    permission_classes = [IsAuthenticated, IsTexnik]
    pagination_class = CustomPagination

    def get_queryset(self):
        user = self.request.user
        qs = KunlikYurish.objects.select_related("tarkib", "created_by")

        if user.role == "texnik" and user.depo:
            qs = qs.filter(tarkib__depo=user.depo)
            
            
        if self.action == "list":
            qs = qs.filter(tarkib__holati="Soz_holatda")

        return qs.order_by("-sana", "-id")

    def perform_create(self, serializer):
        user = self.request.user
        tarkib = serializer.validated_data["tarkib"]

        if user.role == "texnik" and tarkib.depo != user.depo:
            raise PermissionDenied("Siz faqat o‘z depo tarkiblaringiz uchun ma’lumot qo‘shishingiz mumkin.")

        serializer.save(created_by=user)

    @action(detail=True, methods=["get"])
    def total(self, request, pk=None):
        tarkib = self.get_object().tarkib
        total_km = KunlikYurish.objects.filter(tarkib=tarkib).aggregate(Sum("kilometr"))["kilometr__sum"] or 0
        return Response({"tarkib": tarkib.tarkib_raqami, "total_km": total_km})

    @action(detail=False, methods=["get"])
    def by_date(self, request):
        tarkib_id = request.query_params.get("tarkib_id")
        sana = request.query_params.get("sana")  

        if not tarkib_id or not sana:
            return Response({"error": "tarkib_id va sana kerak"}, status=400)

        try:
            sana_date = datetime.strptime(sana, "%d-%m-%Y").date()
        except ValueError:
            return Response({"error": "Sana format noto‘g‘ri, DD-MM-YYYY bo‘lishi kerak"}, status=400)

        qs = KunlikYurish.objects.filter(tarkib_id=tarkib_id)

        daily_km = qs.filter(sana=sana_date).aggregate(Sum("kilometr"))["kilometr__sum"] or 0
        total_until = qs.filter(sana__lte=sana_date).aggregate(Sum("kilometr"))["kilometr__sum"] or 0

        return Response({
            "tarkib_id": tarkib_id,
            "sana": sana,
            "daily_km": daily_km,
            "total_until": total_until
        })

class KunlikYurishHistoryAPIView(APIView):
    permission_classes = [IsAuthenticated, IsTexnik]
    pagination_class = CustomPagination

    def get(self, request, tarkib_id, *args, **kwargs):
        qs = KunlikYurish.objects.filter(
            tarkib_id=tarkib_id
        ).select_related("tarkib", "created_by").order_by("-sana", "-id")

        serializer = KunlikYurishSerializer(qs, many=True, context={"request": request})
        return Response(serializer.data)
    
    @action(detail=False, methods=["get"])
    def by_date(self, request):
        tarkib_id = request.query_params.get("tarkib_id")
        sana = request.query_params.get("sana")  # kutilayotgan format: "DD:MM:YYYY"

        if not tarkib_id or not sana:
            return Response({"error": "tarkib_id va sana kerak"}, status=400)

        try:
            sana_date = datetime.strptime(sana, "%d-%m-%Y").date()
        except ValueError:
            return Response({"error": "Sana format noto‘g‘ri, DD-MM-YYYY bo‘lishi kerak"}, status=400)

        qs = KunlikYurish.objects.filter(tarkib_id=tarkib_id)

        daily_km = qs.filter(sana=sana_date).aggregate(Sum("kilometr"))["kilometr__sum"] or 0
        total_until = qs.filter(sana__lte=sana_date).aggregate(Sum("kilometr"))["kilometr__sum"] or 0

        return Response({
            "tarkib_id": tarkib_id,
            "sana": sana,
            "daily_km": daily_km,
            "total_until": total_until
        })

class TexnikKorikFilter(django_filters.FilterSet):
    tamir_turi_nomi = django_filters.CharFilter(
        field_name="tamir_turi__tamir_nomi", lookup_expr="icontains"
    )
    tarkib_raqami = django_filters.CharFilter(
        field_name="tarkib__tarkib_raqami", lookup_expr="icontains"
    )

    class Meta:
        model = TexnikKorik
        fields = ["tamir_turi_nomi", "tarkib_raqami"]





# --- FAQAT GET ---
class TexnikKorikGetViewSet(mixins.ListModelMixin,
                            mixins.RetrieveModelMixin,
                            viewsets.GenericViewSet):
    queryset = (
        TexnikKorik.objects
        .select_related("tarkib", "tamir_turi", "created_by")
        .prefetch_related(
            "texnikkorikehtiyotqism_set__ehtiyot_qism",  
            "steps__texnikkorikehtiyotqismstep_set__ehtiyot_qism",  
            "steps" 
        )
        .order_by("-id")
    )
    serializer_class = TexnikKorikSerializer
    permission_classes = [IsAuthenticated, IsTexnik]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter, DjangoFilterBackend]
    filterset_class = TexnikKorikFilter
    search_fields = [
        "tarkib__tarkib_raqami",
        "kamchiliklar_haqida",
        "bartaraf_etilgan_kamchiliklar",
        "created_by__username",
    ]
    pagination_class = CustomPagination
    ordering_fields = ["created_at", "approved_at", "kirgan_vaqti"]


    
    def get_queryset(self):
        user = self.request.user
        qs = (
            TexnikKorik.objects
            .select_related("tarkib", "tamir_turi", "created_by")
            .prefetch_related(
                "texnikkorikehtiyotqism_set__ehtiyot_qism",
                "steps__texnikkorikehtiyotqismstep_set__ehtiyot_qism",
                "steps"
            )
            .order_by("-id")
        )

        if user.role == "texnik" and user.depo:
            qs = qs.filter(tarkib__depo=user.depo)
        return qs



class TexnikKorikViewSet(BaseViewSet):
    queryset = TexnikKorik.objects.prefetch_related("steps").all().order_by("-id")
    serializer_class = TexnikKorikSerializer
    permission_classes = [IsAuthenticated, IsTexnik]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter, DjangoFilterBackend]
    search_fields = [
        "tarkib__tarkib_raqami",
        "tamir_turi__tamir_nomi",
        "created_by__username",
        "id",
    ]
    pagination_class = CustomPagination
    
    
    def create(self, request, *args, **kwargs):
        # FormData dan ehtiyot_qismlarni JSON ga o'girish
        if 'ehtiyot_qismlar' in request.data and isinstance(request.data['ehtiyot_qismlar'], str):
            try:
                request.data._mutable = True
                request.data['ehtiyot_qismlar'] = json.loads(request.data['ehtiyot_qismlar'])
                request.data._mutable = False
            except Exception as e:
                print(f"❌ JSON parse xatosi: {e}")
                
        print("YUBORILGAN EHTIYOT QISMLAR:", request.data.get("ehtiyot_qismlar"))

        response = super().create(request, *args, **kwargs)

        # CREATE dan keyin yangi yaratilgan korikni to'liq yuklab olish
        if response.status_code == status.HTTP_201_CREATED:
            korik_id = response.data.get('id')
            if korik_id:
                try:
                    korik = (
                        TexnikKorik.objects
                        .select_related("tarkib", "tamir_turi", "created_by")
                        .prefetch_related(
                            "texnikkorikehtiyotqism_set__ehtiyot_qism",
                            "steps__texnikkorikehtiyotqismstep_set__ehtiyot_qism",
                            "steps"
                        )
                        .get(id=korik_id)
                    )
                    serializer = self.get_serializer(korik)
                    response.data = serializer.data
                except Exception as e:
                    print(f"❌ Korikni yuklab olishda xato: {e}")

        return response
    
    def get_queryset(self):
        user = self.request.user
        qs = TexnikKorik.objects.prefetch_related("steps").order_by("-id")

        if user.role == "texnik" and user.depo:
            qs = qs.filter(tarkib__depo=user.depo)
        return qs

    @extend_schema(
        parameters=[
            OpenApiParameter(name="page", type=int, location=OpenApiParameter.QUERY, description="Step pagination page"),
            OpenApiParameter(name="limit", type=int, location=OpenApiParameter.QUERY, description="Step page size"),
            OpenApiParameter(name="search", type=str, location=OpenApiParameter.QUERY, description="Step search"),
        ]
    )
    def retrieve(self, request, *args, **kwargs):
        """Detail with steps pagination"""
        return super().retrieve(request, *args, **kwargs)
    
    @action(detail=True, methods=["post"], url_path="add-step")
    def add_step(self, request, pk=None):
        korik = self.get_object()
        if korik.status == TexnikKorik.Status.BARTARAF_ETILDI:
            return Response(
                {"detail": "Bu korik yakunlangan, yangi step qo'shib bo'lmaydi!"},
                status=status.HTTP_400_BAD_REQUEST
            )

        serializer = TexnikKorikStepSerializer(
            data=request.data,
            context={"request": request, "korik": korik}  # 👈 korikni contextga qo‘shdik
        )
        serializer.is_valid(raise_exception=True)
        step = serializer.save()  # 👈 korikni save ichida olib beradi
        return Response(TexnikKorikStepSerializer(step).data, status=status.HTTP_201_CREATED)

class TexnikKorikStepViewSet(BaseViewSet):
    serializer_class = TexnikKorikStepSerializer
    permission_classes = [IsAuthenticated, IsTexnik]
    pagination_class = CustomPagination
    filter_backends = [filters.SearchFilter, filters.OrderingFilter, DjangoFilterBackend]

    search_fields = [
        "id",
        "kamchiliklar_haqida",
        "bartaraf_etilgan_kamchiliklar",
        "created_by__username",
        "korik__tarkib__tarkib_raqami",
        "tamir_turi__tamir_nomi",
    ]
    filterset_fields = ["korik"]
    
    

    def get_queryset(self):
        user = self.request.user
        qs = TexnikKorikStep.objects.all().order_by("-id")

        # faqat o‘z depo uchun texnik
        if user.role == "texnik" and user.depo:
            qs = qs.filter(korik__tarkib__depo=user.depo)

        # query param orqali korik_id
        korik_id = self.request.query_params.get("korik")
        if korik_id:
            qs = qs.filter(korik_id=korik_id)

        return qs

    def perform_create(self, serializer):
        korik_id = self.request.query_params.get("korik")
        if not korik_id:
            # agar frontend yubormasa ham, context orqali olishga urinamiz
            korik_id = self.kwargs.get("korik_pk") or self.request.data.get("korik")

        if not korik_id:
            raise ValidationError({"korik": "Texnik korik ID aniqlanmadi!"})

        try:
            korik = TexnikKorik.objects.get(id=korik_id)
        except TexnikKorik.DoesNotExist:
            raise ValidationError({"korik": "Bunday Texnik Korik topilmadi!"})

        if korik.status != TexnikKorik.Status.JARAYONDA:
            raise ValidationError({"korik": "Avval Texnik Korik boshlang yoki u tugallangan."})

        serializer.context["korik"] = korik 
        serializer.save()
        




class NosozliklarFilter(django_filters.FilterSet):
    nosozlik_turi = django_filters.CharFilter(
        field_name="nosozliklar_haqida__nosozlik_turi",
        lookup_expr="icontains"
    )
    tarkib_raqami = django_filters.CharFilter(
        field_name="tarkib__tarkib_raqami",
        lookup_expr="icontains"
    )

    class Meta:
        model = Nosozliklar
        fields = ["nosozlik_turi", "tarkib_raqami", "status"]


class NosozlikStepFilter(django_filters.FilterSet):
    nosozlik_turi = django_filters.CharFilter(
        field_name="nosozlik__nosozliklar_haqida__nosozlik_turi",
        lookup_expr="icontains"
    )
    bartaraf_etilgan_nosozliklar = django_filters.CharFilter(
        lookup_expr="icontains"
    )
    tamir_turi = django_filters.CharFilter(
        field_name="tamir_turi__tamir_nomi",
        lookup_expr="icontains"
    )

    class Meta:
        model = NosozlikStep
        fields = ["nosozlik_turi", "bartaraf_etilgan_nosozliklar", "tamir_turi", "status"]
        
        
class NosozliklarGetViewSet(mixins.ListModelMixin,
                            mixins.RetrieveModelMixin,
                            viewsets.GenericViewSet):
    queryset = (
        Nosozliklar.objects
        .select_related("tarkib", "created_by")
        .prefetch_related("ehtiyot_qismlar")
        .order_by("-id")
    )
    serializer_class = NosozliklarSerializer
    permission_classes = [IsAuthenticated, IsTexnik]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter, DjangoFilterBackend]
    filterset_class = NosozliklarFilter
    search_fields = [
        "nosozliklar_haqida",
        "bartaraf_etilgan_nosozliklar",
        "tarkib__tarkib_raqami",
        "created_by__username",
    ]
    pagination_class = CustomPagination
    ordering_fields = ["created_at", "approved_at", "aniqlangan_vaqti"]


class NosozlikTuriViewSet(viewsets.ModelViewSet):
    queryset = NosozlikTuri.objects.all().order_by("-id")
    serializer_class = NosozlikTuriSerializer
    permission_classes = [IsAuthenticated, IsTexnik]
    
    

class NosozliklarViewSet(BaseViewSet):
    queryset = (
        Nosozliklar.objects
        .select_related("tarkib", "created_by")
        .prefetch_related("steps")
        .order_by("-id")
    )
    serializer_class = NosozliklarSerializer
    permission_classes = [IsAuthenticated, IsTexnik]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter, DjangoFilterBackend]
    filterset_class = NosozliklarFilter
    search_fields = [
        "nosozliklar_haqida__nosozliklar_haqida",
        "bartaraf_etilgan_nosozliklar",
        "tarkib__tarkib_raqami",
        "created_by__username",
        "id",
    ]
    pagination_class = CustomPagination

    def create(self, request, *args, **kwargs):
        # FormData dan ehtiyot qismlarni JSON formatga o‘girish
        if 'ehtiyot_qismlar' in request.data and isinstance(request.data['ehtiyot_qismlar'], str):
            try:
                request.data._mutable = True
                request.data['ehtiyot_qismlar'] = json.loads(request.data['ehtiyot_qismlar'])
                request.data._mutable = False
            except Exception as e:
                print(f"❌ JSON parse xatosi: {e}")

        print("YUBORILGAN NOSOZLIK EHTIYOT QISMLAR:", request.data.get("ehtiyot_qismlar"))

        response = super().create(request, *args, **kwargs)

        # CREATE dan keyin yangilangan obyektni to‘liq qaytarish
        if response.status_code == status.HTTP_201_CREATED:
            nosozlik_id = response.data.get("id")
            if nosozlik_id:
                try:
                    nosozlik = (
                        Nosozliklar.objects
                        .select_related("tarkib", "created_by")
                        .prefetch_related(
                            "ehtiyot_qism_aloqalari__ehtiyot_qism",
                            "steps__ehtiyot_qismlar_step__ehtiyot_qism",
                            "steps"
                        )
                        .get(id=nosozlik_id)
                    )
                    serializer = self.get_serializer(nosozlik)
                    response.data = serializer.data
                except Exception as e:
                    print(f"❌ Nosozlikni yuklashda xato: {e}")

        return response

    def get_queryset(self):
        user = self.request.user
        qs = Nosozliklar.objects.prefetch_related("steps").order_by("-id")
        if user.role == "texnik" and user.depo:
            qs = qs.filter(tarkib__depo=user.depo)
        return qs

    @extend_schema(
        parameters=[
            OpenApiParameter(name="page", type=int, location=OpenApiParameter.QUERY, description="Step pagination page"),
            OpenApiParameter(name="limit", type=int, location=OpenApiParameter.QUERY, description="Step page size"),
            OpenApiParameter(name="search", type=str, location=OpenApiParameter.QUERY, description="Step search"),
        ]
    )
    def retrieve(self, request, *args, **kwargs):
        """Detail with steps pagination"""
        return super().retrieve(request, *args, **kwargs)

    @action(detail=True, methods=["post"], url_path="add-step")
    def add_step(self, request, pk=None):
        nosozlik = self.get_object()
        if nosozlik.status == Nosozliklar.Status.BARTARAF_ETILDI:
            return Response(
                {"detail": "Bu nosozlik yakunlangan, yangi step qo‘shib bo‘lmaydi!"},
                status=status.HTTP_400_BAD_REQUEST
            )

        serializer = NosozlikStepSerializer(
            data=request.data,
            context={"request": request, "nosozlik": nosozlik}
        )
        serializer.is_valid(raise_exception=True)
        step = serializer.save()
        return Response(NosozlikStepSerializer(step).data, status=status.HTTP_201_CREATED)



class NosozlikStepViewSet(BaseViewSet):
    serializer_class = NosozlikStepSerializer
    permission_classes = [IsAuthenticated, IsTexnik]
    pagination_class = CustomPagination
    filter_backends = [filters.SearchFilter, filters.OrderingFilter, DjangoFilterBackend]
    filterset_class = NosozlikStepFilter
    search_fields = [
        "id",
        "nosozliklar_haqida__nosozliklar_haqida",
        "bartaraf_etilgan_nosozliklar",
        "created_by__username",
        "nosozlik__tarkib__tarkib_raqami",
    ]
    filterset_fields = ["nosozlik"]

    def get_queryset(self):
        user = self.request.user
        qs = NosozlikStep.objects.all().order_by("-id")

        if user.role == "texnik" and user.depo:
            qs = qs.filter(nosozlik__tarkib__depo=user.depo)

        nosozlik_id = self.request.query_params.get("nosozlik")
        if nosozlik_id:
            qs = qs.filter(nosozlik_id=nosozlik_id)
        return qs

    def perform_create(self, serializer):
        nosozlik_id = self.request.query_params.get("nosozlik")
        if not nosozlik_id:
            nosozlik_id = self.kwargs.get("nosozlik_pk") or self.request.data.get("nosozlik")

        if not nosozlik_id:
            raise ValidationError({"nosozlik": "Nosozlik ID aniqlanmadi!"})

        try:
            nosozlik = Nosozliklar.objects.get(id=nosozlik_id)
        except Nosozliklar.DoesNotExist:
            raise ValidationError({"nosozlik": "Bunday Nosozlik topilmadi!"})

        if nosozlik.status != Nosozliklar.Status.JARAYONDA:
            raise ValidationError({"nosozlik": "Avval Nosozlik boshlang yoki u yakunlangan."})

        serializer.context["nosozlik"] = nosozlik
        serializer.save()



class NotificationViewSet(viewsets.ModelViewSet):
    serializer_class = NotificationSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        qs = Notification.objects.all().order_by("-created_at")

        # Agar texnik bo‘lsa — faqat o‘ziga tegishli
        if hasattr(user, "role") and user.role == "texnik":
            qs = qs.filter(user=user)
        return qs

    def list(self, request, *args, **kwargs):
        queryset = self.get_queryset()
        nosozliklar = queryset.filter(type="nosozlik")
        texnik_koriklar = queryset.filter(type="texnik_korik")
        ehtiyot_qismlar = queryset.filter(type="ehtiyot_qism")

        return Response({
            "nosozlik": NotificationSerializer(nosozliklar, many=True).data,
            "texnik_korik": NotificationSerializer(texnik_koriklar, many=True).data,
            "ehtiyot_qism": NotificationSerializer(ehtiyot_qismlar, many=True).data,
        })

    @action(detail=True, methods=["patch"])
    def mark_as_read(self, request, pk=None):
        notif = self.get_object()
        notif.is_read = True
        notif.save(update_fields=["is_read"])
        return Response({"detail": "Xabar o‘qilgan sifatida belgilandi"}, status=status.HTTP_200_OK)
        

class KorikNosozlikStatisticsView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        today = now().date()
        first_day_this_month = today.replace(day=1)
        first_day_last_month = (first_day_this_month - timedelta(days=1)).replace(day=1)
        last_day_last_month = first_day_this_month - timedelta(days=1)

        # 1) Umumiy texnik ko‘riklar soni
        total_korik = TexnikKorik.objects.count()

        # 2) Umumiy nosozliklar soni
        total_nosozlik = Nosozliklar.objects.count()

        # 3) Oxirgi oyda bajarilgan texnik ko‘riklar soni
        last_month_korik = TexnikKorik.objects.filter(
            created_at__date__gte=first_day_last_month,
            created_at__date__lte=last_day_last_month
        ).count()

        # ✅ 4) Oxirgi 10 ta nosozlik (eng so‘nggilari)
        last_10_nosozlik = list(
            Nosozliklar.objects.select_related("tarkib")
            .order_by("-created_at")[:10]
            .values(
                "id",
                "tarkib__tarkib_raqami",
                "tarkib__turi",
                "nosozliklar_haqida__nosozlik_turi",  
                "bartaraf_etilgan_nosozliklar",
                "created_at"
                    )
        )

        korik_counts = (
            TexnikKorik.objects.values("tarkib__id", "tarkib__tarkib_raqami", "tarkib__turi")
            .annotate(total=Count("id"))
        )
        nosozlik_counts = (
            Nosozliklar.objects.values("tarkib__id", "tarkib__tarkib_raqami", "tarkib__turi")
            .annotate(total=Count("id"))
        )

        combined = {}
        for item in korik_counts:
            tid = item["tarkib__id"]
            combined[tid] = {
                "id": tid,
                "raqam": item["tarkib__tarkib_raqami"],
                "turi": item["tarkib__turi"],
                "total": combined.get(tid, {}).get("total", 0) + item["total"],
            }

        for item in nosozlik_counts:
            tid = item["tarkib__id"]
            if tid in combined:
                combined[tid]["total"] += item["total"]
            else:
                combined[tid] = {
                    "id": tid,
                    "raqam": item["tarkib__tarkib_raqami"],
                    "turi": item["tarkib__turi"],
                    "total": item["total"],
                }

        top_5_tarkib = sorted(combined.values(), key=lambda x: x["total"], reverse=True)[:5]

        top_10_ehtiyot_qism = list(
        EhtiyotQismlari.objects.annotate(
            qoshilgan=Sum("ehtiyotqism_hist__miqdor"),
            ishlatilgan=(
                Sum("texnikkorikehtiyotqism__miqdor") +
                Sum("texnikkorikehtiyotqismstep__miqdor") +
                Sum("nosozlikehtiyotqism__miqdor") +
                Sum("nosozlikehtiyotqismstep__miqdor")
            )
        )
        .annotate(
            jami_miqdor=F("qoshilgan") - F("ishlatilgan")
        )
        .values("id", "ehtiyotqism_nomi", "birligi", "jami_miqdor")
        .order_by("jami_miqdor")[:10]
    )

        return Response({
            "total_korik": total_korik,
            "total_nosozlik": total_nosozlik,
            "last_month_korik": last_month_korik,
            "last_10_nosozlik": last_10_nosozlik,
            "top_5_tarkib": top_5_tarkib,
            "top_10_ehtiyot_qism": top_10_ehtiyot_qism,
        })




class TarkibDetailViewSet(BaseViewSet):
    permission_classes = [IsAuthenticated, IsTexnik |IsMonitoringReadOnly]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter, DjangoFilterBackend]
    search_fields = ['guruhi','holati']

    def get_queryset(self):
        return HarakatTarkibi.objects.all()

    def get_serializer(self, *args, **kwargs):
        return TarkibFullDetailSerializer(*args, **kwargs)

    def check_permissions(self, request):
        if request.user.is_superuser:
            return
        return super().check_permissions(request)

    def retrieve(self, request, pk=None):
        tarkib = self.get_queryset().filter(pk=pk).first()
        if not tarkib:
            return Response({"detail": "Tarkib topilmadi."}, status=404)

        texnik_koriklar = TexnikKorik.objects.filter(tarkib=tarkib)

        # Barcha tamir turlarini olish
        tamir_turlari = TamirTuri.objects.all().values_list('tamir_nomi', flat=True)

        # Har bir tamir turi uchun 0 dan boshlovchi dict
        tamir_turi_count_map = {t: 0 for t in tamir_turlari}

        # Mavjud texnik koriklar asosida sanash
        for korik in texnik_koriklar:
            if korik.tamir_turi and korik.tamir_turi.tamir_nomi in tamir_turi_count_map:
                tamir_turi_count_map[korik.tamir_turi.tamir_nomi] += 1

        # JSON chiqishi uchun ro‘yxatga aylantirish
        texnik_korik_summary = [
            {"tamir_turi": nomi, "soni": count} for nomi, count in tamir_turi_count_map.items()
        ]

        nosozliklar = Nosozliklar.objects.filter(tarkib=tarkib)
        nosozlik_data = []
        for n in nosozliklar:
            nosozlik_data.append({
                "id": n.id,
                "tarkib": n.tarkib.id if n.tarkib else None,
                "tarkib_nomi": getattr(n.tarkib, "tarkib_raqami", None),
                "nosozlik_turi_id": getattr(n.nosozliklar_haqida, "id", None),
                "nosozlik_turi": getattr(n.nosozliklar_haqida, "nosozlik_turi", None),
                "bartaraf_etilgan_nosozliklar": n.bartaraf_etilgan_nosozliklar,
                "status": n.status,
                "bartarafqilingan_vaqti": n.bartarafqilingan_vaqti,
                "created_by": getattr(n.created_by, "username", None),
                "created_at": n.created_at,
                "akt_file": request.build_absolute_uri(n.akt_file.url) if n.akt_file else None,
            })

        response_data = {
            "tarkib_id": tarkib.id,
            "tarkib_raqami": tarkib.tarkib_raqami,
            "depo_nomi": tarkib.depo.qisqacha_nomi if tarkib.depo else None,
            "guruhi": tarkib.guruhi,
            "turi": tarkib.turi,
            "holati": tarkib.holati,
            "tarkib_photo": request.build_absolute_uri(tarkib.image.url) if getattr(tarkib, 'image', None) else None,
            "ishga_tushgan_vaqti": tarkib.ishga_tushgan_vaqti,
            "eksplutatsiya_vaqti": tarkib.eksplutatsiya_vaqti,
            "created_at": tarkib.created_at,
            "previous_version": tarkib.previous_version.id if getattr(tarkib, 'previous_version', None) else None,
            "created_by": tarkib.created_by.username if getattr(tarkib, 'created_by', None) else None,
            "is_active": tarkib.is_active,
            "texnik_korik_soni_turi": texnik_korik_summary,
            "nosozliklar": nosozlik_data,
        }
        return Response(response_data)

    
    @action(detail=True, methods=["get"], url_path="export-pdf")
    def export_pdf(self, request, pk=None):

        tarkib = self.get_queryset().filter(pk=pk).first()
        if not tarkib:
            return Response({"detail": "Tarkib topilmadi."}, status=404)

        texnik_koriklar = TexnikKorik.objects.filter(tarkib=tarkib)
        nosozliklar = Nosozliklar.objects.filter(tarkib=tarkib).order_by('-id')[:5]

        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="Tarkib_{tarkib.tarkib_raqami}.pdf"'

        buffer = BytesIO()

        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            topMargin=1 * cm,
            bottomMargin=1 * cm,
            leftMargin=1 * cm,
            rightMargin=1 * cm,
        )
        styles = getSampleStyleSheet()
        elements = []

        # --- Sarlavha ---
        title = Paragraph(
            f"<b>{tarkib.tarkib_raqami}</b><br/>"
            f"<b>Harakat tarkibining texnik ko‘rsatkichlari bo‘yicha to‘liq ma’lumot</b>",
            ParagraphStyle('title', parent=styles['Title'], alignment=1, leading=18, textColor=colors.darkblue),
        )
        elements.append(title)
        elements.append(Spacer(1, 10))

        # --- Rasm ---
        if getattr(tarkib, 'image', None):
            try:
                img = Image(tarkib.image.path, width=18.5 * cm, height=9.5 * cm)
                img.hAlign = 'CENTER'
                elements.append(img)
                elements.append(Spacer(1, 10))
            except:
                pass

        # --- Ekspluatatsiya ma’lumotlari ---
        ekspl_data = [
            [
                Paragraph(f"<font color='red'><b>{h}</b></font>", styles['Normal'])
                for h in ["Ekspluatatsiyaga qo‘yilgan sana", "Turi", "Nomeri", "Masofa", "Hozirgi holati"]
            ],
            [
                Paragraph(f"<b>{tarkib.ishga_tushgan_vaqti.strftime('%d-%m-%Y') if tarkib.ishga_tushgan_vaqti else '-'}</b>", 
                            styles['Normal']),
                Paragraph(f"<b>{tarkib.guruhi or '-'}</b>", styles['Normal']),
                Paragraph(f"<b>{tarkib.tarkib_raqami or '-'}</b>", styles['Normal']),
                Paragraph(f"<b>{str(getattr(tarkib, 'masofa', '–'))}</b>", styles['Normal']),
                Paragraph(f"<b>{tarkib.holati or '-'}</b>", styles['Normal']),
            ],
        ]

        # Jadval eni sahifa eni bo‘ylab teng bo‘lishi uchun
        full_width = A4[0] - 2 * cm

        # Ustun kengliklarini belgilash (Nomeri ustunini biroz kengroq qilamiz)
        col_widths = [
            full_width * 0.18,  # Ekspluatatsiya sana
            full_width * 0.18,  # Turi
            full_width * 0.26,  # Nomeri (kattaroq)
            full_width * 0.18,  # Masofa
            full_width * 0.20,  # Hozirgi holati
        ]

        ekspl_table = Table(ekspl_data, hAlign='CENTER', colWidths=col_widths)
        ekspl_table.setStyle(TableStyle([
            ('GRID', (0, 0), (-1, -1), 0.6, colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BACKGROUND', (0, 0), (-1, 0), colors.whitesmoke),
            ('BOX', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
        ]))

        elements.append(ekspl_table)
        elements.append(Spacer(1, 12))

        # --- Texnik ko‘riklar tarixi ---
        elements.append(Paragraph(
            "<b><font color='red' size=12>Texnik ko‘riklar tarixi</font></b>",
            ParagraphStyle('h3', alignment=1)
        ))
        elements.append(Spacer(1, 6))

        # Barcha tamir turlarini bazadan olish
        from home.models import TamirTuri
        tamir_turlari = TamirTuri.objects.all().values_list('tamir_nomi', flat=True)

        # Har bir tamir turi uchun hisob
        korik_counts = {t: 0 for t in tamir_turlari}
        for tk in texnik_koriklar:
            if tk.tamir_turi and tk.tamir_turi.tamir_nomi in korik_counts:
                korik_counts[tk.tamir_turi.tamir_nomi] += 1

        # Jadval uchun ma’lumot tayyorlash
        korik_data = [[Paragraph("<b><font color='black'>Ta’mir turi</font></b>", styles['Normal'])] +
                    [Paragraph(f"<font color='red'><b>{t}</b></font>", styles['Normal']) for t in tamir_turlari]]

        korik_values = [
            Paragraph(f"<b><font color='blue'>{korik_counts[t]}</font></b>", styles['Normal'])
            if korik_counts[t] > 0 else Paragraph("<b>-</b>", styles['Normal'])
            for t in tamir_turlari
        ]

        korik_data.append([Paragraph("<b>Soni</b>", styles['Normal'])] + korik_values)

        # Jadval kengliklarini moslashtirish (dinamik)
        col_width = (A4[0] - 2 * cm) / (len(tamir_turlari) + 1)
        korik_table = Table(korik_data, hAlign='CENTER', colWidths=[col_width] * (len(tamir_turlari) + 1))

        korik_table.setStyle(TableStyle([
            ('GRID', (0, 0), (-1, -1), 0.6, colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BACKGROUND', (0, 0), (-1, 0), colors.whitesmoke),
            ('BOX', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
        ]))
        elements.append(korik_table)
        elements.append(Spacer(1, 12))

        # --- Nosozliklar tarixi ---
        elements.append(Paragraph("<b><font color='red' size=12>Nosozliklar tarixi</font></b>",
                                ParagraphStyle('h3', alignment=1)))
        elements.append(Spacer(1, 6))

        nosoz_data = [[
            Paragraph(f"<font color='darkblue'><b>{h}</b></font>", styles['Normal'])
            for h in ["No", "Nosozlik sababi", "Aniqlangan sana", "Bartarf etilgan sana"]
        ]]

        for i, n in enumerate(nosozliklar, 1):
            aniqlangan = getattr(n, 'aniqlangan_vaqti', None)
            bartaraf = getattr(n, 'bartarafqilingan_vaqti', None)

            # --- Agar asosiy modelda bo'sh bo'lsa, step ichidan eng so'nggisini olamiz ---
            if not bartaraf:
                last_step = n.steps.filter(bartaraf_qilingan_vaqti__isnull=False).order_by('-bartaraf_qilingan_vaqti').first()
                if last_step:
                    bartaraf = last_step.bartaraf_qilingan_vaqti

            # --- sanalarni formatlash ---
            aniqlangan_str = aniqlangan.strftime("%d-%m-%Y %H:%M") if isinstance(aniqlangan, datetime) else "-"
            bartaraf_str = bartaraf.strftime("%d-%m-%Y %H:%M") if isinstance(bartaraf, datetime) else "-"

            nosoz_data.append([
                Paragraph(f"<b>{str(i)}</b>", styles['Normal']),
                Paragraph(f"<b>{getattr(n.nosozliklar_haqida, 'nosozlik_turi', '-') or '-'}</b>", styles['Normal']),
                Paragraph(f"<b>{aniqlangan_str}</b>", styles['Normal']),
                Paragraph(f"<b>{bartaraf_str}</b>", styles['Normal']),
            ])

        nosoz_table = Table(
            nosoz_data,
            hAlign='CENTER',
            colWidths=[1.5 * cm, 10 * cm, 3.5 * cm, 3.5 * cm]
        )
        nosoz_table.setStyle(TableStyle([
            ('GRID', (0, 0), (-1, -1), 0.6, colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BACKGROUND', (0, 0), (-1, 0), colors.whitesmoke),
            ('BOX', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
        ]))
        elements.append(nosoz_table)
        elements.append(Spacer(1, 15))

        # --- QR kod va pastki fon ---
        qr_url = f"https://depo-main.vercel.app/depo/{tarkib.id}"
        qr = qrcode.QRCode(error_correction=qrcode.constants.ERROR_CORRECT_H, box_size=3, border=1)
        qr.add_data(qr_url)
        qr.make(fit=True)
        qr_buffer = BytesIO()
        qr.make_image(fill_color="black", back_color="white").save(qr_buffer, format='PNG')
        qr_buffer.seek(0)
        qr_img = Image(qr_buffer, width=2.5 * cm, height=2.5 * cm)
        qr_img.hAlign = 'RIGHT'

        footer_text = Paragraph(
            """<font size=9><b>
            Hujjatning haqiqiyligini tekshirish uchun QR kodni skanerlang.<br/>
            <font color='blue'><b>depo.tm1.uz</b></font> sayti orqali tasdiqlangan.<br/>
            Ushbu hisobotdagi barcha ma’lumotlar uchun xodim mas’uldir.
            </b></font>""",
            ParagraphStyle('footer', parent=styles['Normal'], alignment=0, leading=12),
        )

        footer_table = Table([[footer_text, qr_img]], colWidths=[14 * cm, 3 * cm])
        footer_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.lightblue),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('BOX', (0, 0), (-1, -1), 0.8, colors.black),
        ]))
        elements.append(footer_table)

        # --- Sahifa border ---
        def draw_border(canvas, doc):
            canvas.saveState()
            canvas.setStrokeColor(colors.black)
            canvas.setLineWidth(1)
            canvas.rect(0.8 * cm, 0.8 * cm, A4[0] - 1.6 * cm, A4[1] - 1.6 * cm)
            canvas.restoreState()

        doc.build(elements, onFirstPage=draw_border, onLaterPages=draw_border)

        buffer.seek(0)
        response.write(buffer.getvalue())
        buffer.close()
        return response







    @action(detail=True, methods=["get"], url_path="export-excel")
    def export_excel(self, request, pk=None):
        tarkib = self.get_queryset().filter(pk=pk).first()
        if not tarkib:
            return Response({"detail": "Tarkib topilmadi."}, status=404)

        texnik_koriklar = TexnikKorik.objects.filter(tarkib=tarkib)
        tamir_turi_count = texnik_koriklar.values('tamir_turi__tamir_nomi').annotate(count=Count('id'))

        nosozliklar = Nosozliklar.objects.filter(tarkib=tarkib)

        wb = Workbook()
        ws = wb.active
        ws.title = f"Tarkib_{tarkib.id}"

        # Sana formatlash yordamchi funksiyasi
        def safe_date(value):
            if hasattr(value, "strftime"):  # agar datetime bo‘lsa
                return value.strftime("%d-%m-%Y")
            return str(value) if value else "-"

        # --- Tarkib umumiy ma'lumotlari ---
        general_info = [
            ["Tarkib ID", tarkib.id],
            ["Tarkib raqami", tarkib.tarkib_raqami],
            ["Depo", tarkib.depo.qisqacha_nomi if tarkib.depo else "-"],
            ["Guruhi", tarkib.guruhi or "-"],
            ["Turi", tarkib.turi or "-"],
            ["Holati", tarkib.holati or "-"],
            ["Ishga tushgan vaqti", safe_date(tarkib.ishga_tushgan_vaqti)],
            ["Eksplutatsiya vaqti", safe_date(tarkib.eksplutatsiya_vaqti)],
            ["Created by", tarkib.created_by.username if tarkib.created_by else "-"],
        ]
        for row in general_info:
            ws.append(row)
        ws.append([])  # bo‘sh qator

        # --- Texnik ko‘riklar jadvali ---
        ws.append(["Texnik ko‘riklar (soni tamir turi bo‘yicha)"])
        ws.append(["Tamir turi", "Soni"])
        for t in tamir_turi_count:
            ws.append([t.get("tamir_turi__tamir_nomi", "-"), t.get("count", 0)])
        ws.append([])

        # --- Nosozliklar jadvali ---
        ws.append(["Nosozliklar batafsil"])
        ws.append(["ID", "Nosozlik sababi", "Bartarf etilgan nosozlik", "Holati"])

        for n in nosozliklar:
            nosozlik_sababi = "-"
            if hasattr(n, "nosozliklar_haqida") and n.nosozliklar_haqida:
                nosozlik_sababi = getattr(n.nosozliklar_haqida, "nosozlik_turi", str(n.nosozliklar_haqida))
            bartaraf = getattr(n, "bartaraf_etilgan_nosozliklar", "-") or "-"
            holat = getattr(n, "status", "-") or "-"
            ws.append([n.id, str(nosozlik_sababi), str(bartaraf), str(holat)])

        # --- Foydalanuvchiga qaytarish ---
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="Tarkib_{tarkib.id}.xlsx"'

        wb.save(response)
        return response

   
   
   

class TexnikKorikByTypeViewSet(BaseViewSet):
    queryset = TexnikKorik.objects.select_related("tarkib", "tamir_turi", "created_by")
    serializer_class = TexnikKorikSerializer
    permission_classes = [IsAuthenticated, IsTexnik |IsMonitoringReadOnly]

    # -------- PDF EXPORT --------
    @action(detail=False, methods=["get"], url_path=r"(?P<tarkib_id>\d+)/(?P<tamir_turi_id>\d+)/export-pdf")
    def export_pdf(self, request, tarkib_id=None, tamir_turi_id=None):
        queryset = self.get_queryset().filter(tarkib_id=tarkib_id, tamir_turi_id=tamir_turi_id)
        if not queryset.exists():
            return HttpResponse("Ma'lumot topilmadi", status=404)

        tarkib = queryset.first().tarkib
        tamir_turi = queryset.first().tamir_turi

        buffer = BytesIO()
        response = HttpResponse(content_type="application/pdf")
        filename = f"{tarkib.tarkib_raqami}_{tamir_turi.tamir_nomi}_koriklar.pdf"
        response["Content-Disposition"] = f'attachment; filename="{filename}"'

        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=1*cm, bottomMargin=1*cm)
        styles = getSampleStyleSheet()
        elements = []
        title_text = (
            f"<b><font color='red'>{tarkib.tarkib_raqami}</font></b> harakat tarkibining "
            f"<br/><b><font color='red'>{tamir_turi.tamir_nomi}</font></b> texnik ko'rigi bo‘yicha to'liq ma'lumot"
        )
        # --- Sarlavha ---
        title = Paragraph(
            title_text,
            ParagraphStyle('title', parent=styles['Title'], alignment=1, textColor=colors.darkblue),
        )
        elements.append(title)
        elements.append(Spacer(1, 10))

        # --- Rasm ---
        if getattr(tarkib, 'image', None):
            try:
                img = Image(tarkib.image.path, width=15.5 * cm, height=6.5 * cm)
                img.hAlign = 'CENTER'
                elements.append(img)
                elements.append(Spacer(1, 10))
            except:
                pass

        # --- Ekspluatatsiya ma’lumotlari ---
        ekspl_data = [
            [
                Paragraph(f"<font color='red'><b>{h}</b></font>", styles['Normal'])
                for h in ["Ekspluatatsiyaga qo‘yilgan sana", "Turi", "Nomeri", "Masofa", "Hozirgi holati"]
            ],
            [
                Paragraph(f"<b>{tarkib.ishga_tushgan_vaqti.strftime('%d-%m-%Y') if tarkib.ishga_tushgan_vaqti else '-'}</b>", 
                            styles['Normal']),
                Paragraph(f"<b>{tarkib.guruhi or '-'}</b>", styles['Normal']),
                Paragraph(f"<b>{tarkib.tarkib_raqami or '-'}</b>", styles['Normal']),
                Paragraph(f"<b>{str(getattr(tarkib, 'masofa', '–'))}</b>", styles['Normal']),
                Paragraph(f"<b>{tarkib.holati or '-'}</b>", styles['Normal']),
            ],
        ]

        # Jadval eni sahifa eni bo‘ylab teng bo‘lishi uchun
        full_width = A4[0] - 2 * cm

        # Ustun kengliklarini belgilash (Nomeri ustunini biroz kengroq qilamiz)
        col_widths = [
            full_width * 0.18,  # Ekspluatatsiya sana
            full_width * 0.18,  # Turi
            full_width * 0.26,  # Nomeri (kattaroq)
            full_width * 0.18,  # Masofa
            full_width * 0.20,  # Hozirgi holati
        ]

        ekspl_table = Table(ekspl_data, hAlign='CENTER', colWidths=col_widths)
        ekspl_table.setStyle(TableStyle([
            ('GRID', (0, 0), (-1, -1), 0.6, colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BACKGROUND', (0, 0), (-1, 0), colors.whitesmoke),
            ('BOX', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
        ]))

        elements.append(ekspl_table)
        elements.append(Spacer(1, 12))

        # --- Texnik ko‘riklar tarixi (Soddalashtirilgan) ---
        history_title = Paragraph(
            f"<b><font color='red'>{tamir_turi.tamir_nomi} texnik ko'rigi tarixi</font></b>",
            ParagraphStyle('h3', alignment=1)
        )
        elements.append(history_title)
        elements.append(Spacer(1, 6))

        data = [[
            Paragraph(f"<font color='darkblue'><b>{h}</b></font>", styles['Normal'])
            for h in ["No", "Kirgan sana", "Chiqqan sana", "Kim (kimlar) tomonidan  texnik ko’rik o’tkazilganligi"]
        ]]

        for i, k in enumerate(queryset, 1):
            # --- Chiqqan vaqtni aniqlaymiz ---
            chiqqan = getattr(k, 'chiqqan_vaqti', None)
            
            # Agar chiqqan_vaqti bo'sh bo'lsa, step ichidan eng so'nggisini olamiz
            if not chiqqan:
                last_step = k.steps.filter(chiqqan_vaqti__isnull=False).order_by('-chiqqan_vaqti').first()
                if last_step:
                    chiqqan = last_step.chiqqan_vaqti

            # --- 🔹 Foydalanuvchilarni jamlash (asosiy + step ichidagilar) ---
            users = set()

            if getattr(k, "created_by", None):
                users.add(k.created_by.username)
            if getattr(k, "step_yozgan", None):
                users.add(k.step_yozgan.username)
            if getattr(k, "yakunlagan", None):
                users.add(k.yakunlagan.username)

            # Step ichidagi foydalanuvchilar
            step_users = k.steps.values_list("created_by__username", flat=True)
            users.update(u for u in step_users if u)

            users_str = ", ".join(sorted(users)) if users else "-"

            # Sana formatlari
            kirgan_str = k.kirgan_vaqti.strftime("%d-%m-%Y %H:%M") if k.kirgan_vaqti else "-"
            chiqqan_str = chiqqan.strftime("%d-%m-%Y %H:%M") if chiqqan else "-"

            data.append([
                i,
                kirgan_str,
                chiqqan_str,
                users_str
            ])

        # Jadval ustun kengliklari
        col_widths_history = [full_width * 0.1, full_width * 0.3, full_width * 0.3, full_width * 0.3]

        history_table = Table(data, hAlign='CENTER', colWidths=col_widths_history)
        history_table.setStyle(TableStyle([
            ('GRID', (0,0), (-1,-1), 0.6, colors.black),
            ('BACKGROUND', (0,0), (-1,0), colors.whitesmoke),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ]))
        elements.append(history_table)
        elements.append(Spacer(1, 12))


        # --- QR kod ---
        qr_url = f"https://depo-main.vercel.app/depo/{tarkib.id}"
        qr = qrcode.QRCode(error_correction=qrcode.constants.ERROR_CORRECT_H, box_size=3, border=1)
        qr.add_data(qr_url)
        qr.make(fit=True)
        qr_buffer = BytesIO()
        qr.make_image(fill_color="black", back_color="white").save(qr_buffer, format='PNG')
        qr_buffer.seek(0)
        qr_img = Image(qr_buffer, width=2.5 * cm, height=2.5 * cm)
        qr_img.hAlign = 'RIGHT'

        footer_text = Paragraph(
            """<font size=9><b>
            Hujjatning haqiqiyligini tekshirish uchun QR kodni skanerlang.<br/>
            <font color='blue'><b>depo.tm1.uz</b></font> sayti orqali tasdiqlangan.<br/>
            Ushbu hisobotdagi barcha ma’lumotlar uchun xodim mas’uldir.
            </b></font>""",
            ParagraphStyle('footer', parent=styles['Normal'], alignment=0, leading=12),
        )

        footer_table = Table([[footer_text, qr_img]], colWidths=[14 * cm, 3 * cm])
        footer_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.lightblue),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('BOX', (0, 0), (-1, -1), 0.8, colors.black),
        ]))
        elements.append(footer_table)

        # --- Sahifa border ---
        def draw_border(canvas, doc):
            canvas.saveState()
            canvas.setStrokeColor(colors.black)
            canvas.setLineWidth(1)
            canvas.rect(0.8 * cm, 0.8 * cm, A4[0] - 1.6 * cm, A4[1] - 1.6 * cm)
            canvas.restoreState()

        doc.build(elements, onFirstPage=draw_border, onLaterPages=draw_border)

        buffer.seek(0)
        response.write(buffer.getvalue())
        buffer.close()
        return response


    # -------- EXCEL EXPORT --------
    @action(detail=False, methods=["get"], url_path=r"(?P<tarkib_id>\d+)/(?P<tamir_turi_id>\d+)/export-excel")
    def export_excel(self, request, tarkib_id=None, tamir_turi_id=None):
        queryset = self.get_queryset().filter(tarkib_id=tarkib_id, tamir_turi_id=tamir_turi_id)
        if not queryset.exists():
            return HttpResponse("Ma'lumot topilmadi", status=404)

        tarkib = queryset.first().tarkib
        tamir_turi = queryset.first().tamir_turi

        wb = Workbook()
        ws = wb.active
        ws.title = "Texnik Ko'riklar"

        ws.merge_cells("A1:G1")
        ws["A1"] = f"{tarkib.tarkib_raqami} — {tamir_turi.tamir_nomi} bo‘yicha texnik ko‘riklar"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A1"].alignment = Alignment(horizontal="center")

        ws.append([])
        ws.append(["Ekspluatatsiya", "Turi", "Raqami", "Masofa", "Holati"])
        ws.append([
            str(tarkib.eksplutatsiya_vaqti or "-"),
            tarkib.turi or "-",
            tarkib.tarkib_raqami or "-",
            getattr(tarkib, "masofa", "-"),
            tarkib.holati or "-",
        ])
        ws.append([])

        ws.append(["#", "Kirgan sana", "Chiqqan sana", "Status", "Kamchiliklar", "Bart. etilgan", "Yaratgan xodim"])
        for i, k in enumerate(queryset, 1):
            ws.append([
                i,
                str(k.kirgan_vaqti or "-"),
                str(k.chiqqan_vaqti or "-"),
                k.status or "-",
                k.kamchiliklar_haqida or "-",
                k.bartaraf_etilgan_kamchiliklar or "-",
                k.created_by.username if k.created_by else "-",
            ])

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        filename = f"{tarkib.tarkib_raqami}_{tamir_turi.tamir_nomi}_koriklar.xlsx"
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response





class TexnikKorikStepViewSet1(ViewSet):
    permission_classes = [IsAuthenticated, IsTexnik |IsMonitoringReadOnly]

    def get_queryset(self):
        return TexnikKorikStep.objects.select_related(
            "korik__tarkib", "tamir_turi", "created_by"
        ).prefetch_related("ehtiyot_qismlar").all()

    @action(detail=False, methods=["get"], url_path=r"(?P<korik_id>\d+)/export-pdf")
    def export_pdf(self, request, korik_id=None):
        # --- Korikni olish ---
        try:
            korik = TexnikKorik.objects.select_related("tarkib", "tamir_turi").get(id=korik_id)
        except TexnikKorik.DoesNotExist:
            return HttpResponse("Bunday Texnik Ko‘rik topilmadi", status=404)

        # --- Shu korik bo‘yicha barcha steplar ---
        steps = self.get_queryset().filter(korik=korik).order_by("id")
        if not steps.exists():
            return HttpResponse("Bu ko‘rik bo‘yicha steplar topilmadi", status=404)

        buffer = BytesIO()
        response = HttpResponse(content_type="application/pdf")
        filename = f"Korik_{korik.id}_steps.pdf"
        response["Content-Disposition"] = f'attachment; filename="{filename}"'

        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=1*cm, bottomMargin=1*cm)
        styles = getSampleStyleSheet()
        elements = []

        # --- Title tepa ---
        title_text = (
            f"<b><font color='red'>{korik.tarkib.tarkib_raqami if korik.tarkib else '-'}</font></b> harakat tarkibining "
            f"<b><font color='red'>{korik.created_at.strftime('%d-%m-%Y') if korik.created_at else '-'}</font></b> da o'tkazilgan "
            f"<b><font color='red'>{korik.tamir_turi.tamir_nomi if korik.tamir_turi else '-'}</font></b> texnik ko‘rigi bo'yicha to'liq ma'lumot"
        )
        title = Paragraph(
            title_text, 
            ParagraphStyle('title', parent=styles['Title'], alignment=1, textColor=colors.darkblue)
        )
        elements.append(title)
        elements.append(Spacer(1, 10))

        # --- Tarkib rasmi ---
        if getattr(korik.tarkib, 'image', None):
            try:
                img = Image(korik.tarkib.image.path, width=15.5*cm, height=6.5*cm)
                img.hAlign = 'CENTER'
                elements.append(img)
                elements.append(Spacer(1, 10))
            except:
                pass

        # --- Ekspluatatsiya jadvali ---
        ekspl_data = [
            [
                Paragraph(f"<font color='red'><b>{h}</b></font>", styles['Normal'])
                for h in ["Ekspluatatsiyaga qo‘yilgan sana", "Turi", "Nomeri", "Masofa", "Hozirgi holati"]
            ],
            [
                Paragraph(f"<b>{korik.tarkib.ishga_tushgan_vaqti.strftime('%d-%m-%Y') if getattr(korik.tarkib, 'ishga_tushgan_vaqti', None) else '-'}</b>", styles['Normal']),
                Paragraph(f"<b>{korik.tarkib.guruhi if getattr(korik.tarkib, 'guruhi', None) else '-'}</b>", styles['Normal']),
                Paragraph(f"<b>{korik.tarkib.tarkib_raqami if getattr(korik.tarkib, 'tarkib_raqami', None) else '-'}</b>", styles['Normal']),
                Paragraph(f"<b>{korik.tarkib.masofa if getattr(korik.tarkib, 'masofa', None) else '-'}</b>", styles['Normal']),
                Paragraph(f"<b>{korik.tarkib.holati if getattr(korik.tarkib, 'holati', None) else '-'}</b>", styles['Normal']),
            ],
        ]
        full_width = A4[0]-2*cm
        col_widths = [full_width*0.18, full_width*0.18, full_width*0.26, full_width*0.18, full_width*0.20]

        ekspl_table = Table(ekspl_data, hAlign='CENTER', colWidths=col_widths)
        ekspl_table.setStyle(TableStyle([
            ('GRID', (0,0), (-1,-1), 0.6, colors.black),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('BACKGROUND', (0,0), (-1,0), colors.whitesmoke),
            ('BOX', (0,0), (-1,-1), 1, colors.black),
            ('FONTSIZE', (0,0), (-1,-1), 9),
        ]))
        elements.append(ekspl_table)
        elements.append(Spacer(1, 12))

        # --- Texnik ko‘rik steplari ---
        created_at_str = korik.created_at.strftime('%d-%m-%Y') if getattr(korik, 'created_at', None) else '-'
        tamir_turi_nomi = getattr(korik.tamir_turi, 'tamir_nomi', '-') if getattr(korik, 'tamir_turi', None) else '-'
        tarkib_raqami = getattr(korik.tarkib, 'tarkib_raqami', '-') if getattr(korik, 'tarkib', None) else '-'

        history_title_text = f"<font color='red'>{created_at_str}</font> da o'tkazilgan " \
                            f"<font color='red'>{tamir_turi_nomi}</font> texnik ko'rigi bo'yicha to'liq ma'lumot"
        history_title = Paragraph(
            history_title_text,
            ParagraphStyle('h3', alignment=1, leading=14, textColor=colors.darkblue)
        )
        elements.append(history_title)
        elements.append(Spacer(1, 6))

        # --- Qizil chiziq Title dan keyin ---
        title_line = Table(
            [[""]],
            colWidths=[A4[0] - 2 * cm],
            style=TableStyle([('LINEABOVE', (0,0), (-1,-1), 1, colors.red)])
        )
        elements.append(title_line)
        elements.append(Spacer(1, 6))

        # --- Steplar ---
        for i, step in enumerate(steps, 1):
            created_by = getattr(step.created_by, 'username', '-')
            kamchilik = getattr(step, "kamchiliklar_haqida", '-')

            # ✅ To‘g‘ri joy: through model orqali olish
            ehtiyot_qismlar_detail = ", ".join([
                f"{eqs.ehtiyot_qism.ehtiyotqism_nomi} ({eqs.miqdor} {eqs.ehtiyot_qism.birligi})"
                for eqs in step.texnikkorikehtiyotqismstep_set.select_related("ehtiyot_qism").all()
            ]) or "-"

            step_date = step.created_at.strftime("%d-%m-%Y") if getattr(step, 'created_at', None) else "-"

            step_text = (
                f"<font color='darkblue'>{tarkib_raqami} harakat tarkibi "
                f"<font color='red'>{step_date}</font> da "
                f"<font color='red'>{created_by}</font> Depo navbatchisi tomonidan "
                f"<font color='red'>{tamir_turi_nomi}</font> texnik ko'rigiga ro'yhatga olindi.<br/><br/>"
                f"Depo navbatchisi aniqlagan kamchiliklar: <font color='red'>{kamchilik}</font><br/>"
                f"Depo navbatchisi ishlatgan ehtiyot qismlar: <font color='red'>{ehtiyot_qismlar_detail}</font>"
                f"</font>"
            )

            p = Paragraph(step_text, ParagraphStyle('Normal', alignment=1, leading=12))
            elements.append(p)
            elements.append(Spacer(1, 6))

            # --- Chiziq Step dan keyin ---
            step_line = Table(
                [[""]],
                colWidths=[A4[0] - 2 * cm],
                style=TableStyle([('LINEABOVE', (0,0), (-1,-1), 1, colors.red)])
            )
            elements.append(step_line)
            elements.append(Spacer(1, 6))

        # --- Yakuniy matn: AKT ---
        akt_text = f"<font color='darkblue'>Tamir turi <font color='red'>{tamir_turi_nomi}</font> to'liq yakunlanganligi bo'yicha AKT ilova qilindi.</font>"
        akt_para = Paragraph(
            akt_text,
            ParagraphStyle('Normal', alignment=1, leading=12)
        )
        elements.append(Spacer(1, 12))
        elements.append(akt_para)
        elements.append(Spacer(1, 6))



        # --- Footer va QR ---
        qr_url = f"https://depo-main.vercel.app/depo/{korik.tarkib.id if korik.tarkib else 0}"
        qr = qrcode.QRCode(error_correction=qrcode.constants.ERROR_CORRECT_H, box_size=3, border=1)
        qr.add_data(qr_url)
        qr.make(fit=True)
        qr_buffer = BytesIO()
        qr.make_image(fill_color="black", back_color="white").save(qr_buffer, format='PNG')
        qr_buffer.seek(0)
        qr_img = Image(qr_buffer, width=2.5*cm, height=2.5*cm)
        qr_img.hAlign = 'RIGHT'

        footer_text = Paragraph(
            """<font size=9><b>
            Hujjatning haqiqiyligini tekshirish uchun QR kodni skanerlang.<br/>
            <font color='blue'><b>depo.tm1.uz</b></font> sayti orqali tasdiqlangan.<br/>
            Ushbu hisobotdagi barcha ma’lumotlar uchun xodim mas’uldir.
            </b></font>""",
            ParagraphStyle('footer', parent=styles['Normal'], alignment=0, leading=12),
        )
        footer_table = Table([[footer_text, qr_img]], colWidths=[14*cm, 3*cm])
        footer_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,-1), colors.lightblue),
            ('VALIGN', (0,0), (-1,-1), 'TOP'),
            ('BOX', (0,0), (-1,-1), 0.8, colors.black),
        ]))
        elements.append(footer_table)

        # --- Border ---
        def draw_border(canvas, doc):
            canvas.saveState()
            canvas.setStrokeColor(colors.black)
            canvas.setLineWidth(1)
            canvas.rect(0.8*cm, 0.8*cm, A4[0]-1.6*cm, A4[1]-1.6*cm)
            canvas.restoreState()

        doc.build(elements, onFirstPage=draw_border, onLaterPages=draw_border)
        buffer.seek(0)
        response.write(buffer.getvalue())
        buffer.close()
        return response


    @action(detail=True, methods=["get"], url_path="export-excel")
    def export_step_excel(self, request, pk=None):
        try:
            korik = TexnikKorik.objects.get(id=pk)
        except TexnikKorik.DoesNotExist:
            return HttpResponse("Bunday Texnik Korik topilmadi", status=404)

        steps = self.get_queryset().filter(korik=korik).order_by("id")
        if not steps.exists():
            return HttpResponse("Bu ko'rik bo'yicha steplar topilmadi", status=404)

        wb = Workbook()
        ws = wb.active
        ws.title = "Texnik Ko'rik Steplar"

        ws.append(["#", "Tarkib raqami", "Tamir turi", "Kamchiliklar", "Ehtiyot qismlar", "Step yozgan", "Sana"])

        for i, step in enumerate(steps, 1):
            tarkib_raqami = getattr(getattr(step.korik, 'tarkib', None), 'tarkib_raqami', '-')
            tamir_turi = getattr(step.tamir_turi, 'tamir_nomi', '-')
            kamchilik = step.kamchiliklar_haqida or "-"
            ehtiyot_qismlar_detail = ", ".join([getattr(eq, 'ehtiyot_qism_nomi', '-') for eq in step.ehtiyot_qismlar.all()]) or "-"
            created_by = getattr(step.created_by, 'username', '-')
            step_date = step.created_at.strftime("%d-%m-%Y") if step.created_at else "-"

            ws.append([
                i,
                tarkib_raqami,
                tamir_turi,
                kamchilik,
                ehtiyot_qismlar_detail,
                created_by,
                step_date
            ])

        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = f'attachment; filename="Korik_{korik.id}_steps.xlsx"'
        wb.save(response)
        return response

    
    
class NosozliklarPDFExportView(ViewSet):
    permission_classes = [IsAuthenticated, IsTexnik |IsMonitoringReadOnly]

    def get_queryset(self):
        return Nosozliklar.objects.all()
    
    @action(detail=True, methods=["get"], url_path="export-pdf")
    def export_pdf(self, request, pk=None):
        tarkib = HarakatTarkibi.objects.filter(pk=pk).first()
        if not tarkib:
            return HttpResponse("Tarkib topilmadi", status=404)

        nosozliklar = Nosozliklar.objects.filter(tarkib=tarkib).order_by('-id')
        if not nosozliklar.exists():
            return HttpResponse("Ushbu tarkibda nosozliklar topilmadi", status=404)

        buffer = BytesIO()
        response = HttpResponse(content_type="application/pdf")
        filename = f"Tarkib_{tarkib.tarkib_raqami}_nosozliklar.pdf"
        response["Content-Disposition"] = f'attachment; filename="{filename}"'

        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=1*cm, bottomMargin=1*cm)
        styles = getSampleStyleSheet()
        elements = []

        # Sarlavha
        title_text = (
            f"<b><font color='red'>{tarkib.tarkib_raqami}</font></b> harakat tarkibining "
            f"<br/><b><font color='red'>Nosozliklar tarixi bo‘yicha to‘liq ma’lumot</font></b>"
        )
        title = Paragraph(title_text, ParagraphStyle('title', parent=styles['Title'], alignment=1, textColor=colors.darkblue))
        elements.append(title)
        elements.append(Spacer(1, 10))

        # --- Tarkib rasmi ---
        if getattr(tarkib, 'image', None):
            try:
                img = Image(tarkib.image.path, width=16 * cm, height=7 * cm)
                img.hAlign = 'CENTER'
                elements.append(img)
                elements.append(Spacer(1, 10))
            except:
                pass

        # --- Tarkib haqida ma’lumot ---
        ekspl_data = [
            [
                Paragraph(f"<font color='red'><b>{h}</b></font>", styles['Normal'])
                for h in ["Ekspluatatsiyaga qo‘yilgan sana", "Turi", "Nomeri", "Masofa", "Hozirgi holati"]
            ],
            [
                Paragraph(f"<b>{tarkib.ishga_tushgan_vaqti.strftime('%d-%m-%Y') if tarkib.ishga_tushgan_vaqti else '-'}</b>", styles['Normal']),
                Paragraph(f"<b>{tarkib.guruhi or '-'}</b>", styles['Normal']),
                Paragraph(f"<b>{tarkib.tarkib_raqami or '-'}</b>", styles['Normal']),
                Paragraph(f"<b>{str(getattr(tarkib, 'masofa', '–'))}</b>", styles['Normal']),
                Paragraph(f"<b>{tarkib.holati or '-'}</b>", styles['Normal']),
            ],
        ]

        full_width = A4[0] - 2 * cm
        col_widths = [full_width * 0.18, full_width * 0.18, full_width * 0.26, full_width * 0.18, full_width * 0.20]
        ekspl_table = Table(ekspl_data, hAlign='CENTER', colWidths=col_widths)
        ekspl_table.setStyle(TableStyle([
            ('GRID', (0, 0), (-1, -1), 0.6, colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BACKGROUND', (0, 0), (-1, 0), colors.whitesmoke),
            ('BOX', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
        ]))
        elements.append(ekspl_table)
        elements.append(Spacer(1, 12))

        # --- Nosozliklar tarixi ---
        elements.append(Paragraph("<b><font color='red' size=12>Nosozliklar tarixi</font></b>", ParagraphStyle('h3', alignment=1)))
        elements.append(Spacer(1, 6))

        nosoz_data = [[
            Paragraph(f"<font color='darkblue'><b>{h}</b></font>", styles['Normal'])
            for h in ["No", "Nosozlik sababi", "Aniqlangan sana", "Bartarf etilgan sana", "Kimlar qilgani"]
        ]]

        for i, n in enumerate(nosozliklar, 1):
            aniqlangan = getattr(n, 'aniqlangan_vaqti', None)
            bartaraf = getattr(n, 'bartarafqilingan_vaqti', None)
            if not bartaraf:
                last_step = n.steps.filter(bartaraf_qilingan_vaqti__isnull=False).order_by('-bartaraf_qilingan_vaqti').first()
                if last_step:
                    bartaraf = last_step.bartaraf_qilingan_vaqti

            aniqlangan_str = aniqlangan.strftime("%d-%m-%Y %H:%M") if isinstance(aniqlangan, datetime) else "-"
            bartaraf_str = bartaraf.strftime("%d-%m-%Y %H:%M") if isinstance(bartaraf, datetime) else "-"

            # 🔹 Kimlar qilgani
            users = set()
            if getattr(n, "created_by", None):
                users.add(n.created_by.username)
            if getattr(n, "yakunlagan", None):
                users.add(n.yakunlagan.username)
            if getattr(n, "step_yozgan", None):
                users.add(n.step_yozgan.username)

            step_users = n.steps.values_list("created_by__username", flat=True)
            users.update(u for u in step_users if u)
            users_str = ", ".join(sorted(users)) if users else "-"

            # 🔹 Nosozlik sababi (asosiy + step ichidagilar)
            sabablar = set()
            if getattr(n.nosozliklar_haqida, "nosozlik_turi", None):
                sabablar.add(n.nosozliklar_haqida.nosozlik_turi)

            # Step ichidagi barcha nosozlik sabablarini yig‘ish
            step_sabablar = n.steps.values_list("nosozliklar_haqida__nosozlik_turi", flat=True)
            sabablar.update(s for s in step_sabablar if s)
            sabablar_str = ", ".join(sorted(sabablar)) if sabablar else "-"

            nosoz_data.append([
                Paragraph(f"<b>{i}</b>", styles['Normal']),
                Paragraph(f"<b>{sabablar_str}</b>", styles['Normal']),
                Paragraph(f"<b>{aniqlangan_str}</b>", styles['Normal']),
                Paragraph(f"<b>{bartaraf_str}</b>", styles['Normal']),
                Paragraph(f"<b>{users_str}</b>", styles['Normal']),
            ])

        nosoz_table = Table(
            nosoz_data,
            hAlign='CENTER',
            colWidths=[1.2 * cm, 6.5 * cm, 3.2 * cm, 3.2 * cm, 5.0 * cm]
        )
        nosoz_table.setStyle(TableStyle([
            ('GRID', (0, 0), (-1, -1), 0.6, colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BACKGROUND', (0, 0), (-1, 0), colors.whitesmoke),
            ('BOX', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
        ]))
        elements.append(nosoz_table)
        elements.append(Spacer(1, 15))

        # --- QR kod va footer ---
        qr_url = f"https://depo-main.vercel.app/depo/{tarkib.id}"
        qr = qrcode.QRCode(error_correction=qrcode.constants.ERROR_CORRECT_H, box_size=3, border=1)
        qr.add_data(qr_url)
        qr.make(fit=True)
        qr_buffer = BytesIO()
        qr.make_image(fill_color="black", back_color="white").save(qr_buffer, format='PNG')
        qr_buffer.seek(0)
        qr_img = Image(qr_buffer, width=2.5 * cm, height=2.5 * cm)
        qr_img.hAlign = 'RIGHT'

        footer_text = Paragraph(
            """<font size=9><b>
            Hujjatning haqiqiyligini tekshirish uchun QR kodni skanerlang.<br/>
            <font color='blue'><b>depo.tm1.uz</b></font> sayti orqali tasdiqlangan.<br/>
            Ushbu hisobotdagi barcha ma’lumotlar uchun xodim mas’uldir.
            </b></font>""",
            ParagraphStyle('footer', parent=styles['Normal'], alignment=0, leading=12),
        )

        footer_table = Table([[footer_text, qr_img]], colWidths=[14 * cm, 3 * cm])
        footer_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.lightblue),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('BOX', (0, 0), (-1, -1), 0.8, colors.black),
        ]))
        elements.append(footer_table)

        def draw_border(canvas, doc):
            canvas.saveState()
            canvas.setStrokeColor(colors.black)
            canvas.setLineWidth(1)
            canvas.rect(0.8 * cm, 0.8 * cm, A4[0] - 1.6 * cm, A4[1] - 1.6 * cm)
            canvas.restoreState()

        doc.build(elements, onFirstPage=draw_border, onLaterPages=draw_border)
        buffer.seek(0)
        response.write(buffer.getvalue())
        buffer.close()
        return response

    # ======================
    # 🔹 EXCEL EXPORT ACTION
    # ======================
    @action(detail=True, methods=["get"], url_path="export-excel")
    def export_excel(self, request, pk=None):
        tarkib = HarakatTarkibi.objects.filter(pk=pk).first()
        if not tarkib:
            return HttpResponse("Tarkib topilmadi", status=404)

        nosozliklar = Nosozliklar.objects.filter(tarkib=tarkib).order_by('-id')
        if not nosozliklar.exists():
            return HttpResponse("Ushbu tarkibda nosozliklar topilmadi", status=404)

        wb = Workbook()
        ws = wb.active
        ws.title = "Nosozliklar tarixi"

        ws.append(["#", "Nosozlik sababi", "Aniqlangan sana", "Bartarf etilgan sana"])

        for i, n in enumerate(nosozliklar, 1):
            aniqlangan = getattr(n, 'aniqlangan_vaqti', None)
            bartaraf = getattr(n, 'bartarafqilingan_vaqti', None)
            aniqlangan_str = aniqlangan.strftime("%d-%m-%Y %H:%M") if isinstance(aniqlangan, datetime) else "-"
            bartaraf_str = bartaraf.strftime("%d-%m-%Y %H:%M") if isinstance(bartaraf, datetime) else "-"
            ws.append([
                i,
                getattr(n.nosozliklar_haqida, 'nosozlik_turi', '-') or "-",
                aniqlangan_str,
                bartaraf_str,
            ])

        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = f'attachment; filename="Tarkib_{tarkib.tarkib_raqami}_nosozliklar.xlsx"'
        wb.save(response)
        return response
    



class NosozlikStepViewSet1(ViewSet):
    permission_classes = [IsAuthenticated, IsTexnik |IsMonitoringReadOnly]

    def get_queryset(self):
        return NosozlikStep.objects.select_related(
            "nosozlik__tarkib", "nosozlik__nosozliklar_haqida", "created_by"
        ).prefetch_related("ehtiyot_qismlar_step__ehtiyot_qism")

    # === PDF EXPORT ===
    @action(detail=True, methods=["get"], url_path="export-pdf")
    def export_pdf(self, request, pk=None):
        try:
            nosozlik = Nosozliklar.objects.select_related("tarkib", "nosozliklar_haqida").get(id=pk)
        except Nosozliklar.DoesNotExist:
            return HttpResponse("Bunday nosozlik topilmadi", status=404)

        steps = self.get_queryset().filter(nosozlik_id=pk).order_by("id")
        if not steps.exists():
            return HttpResponse("Bu nosozlik bo‘yicha steplar topilmadi", status=404)

        buffer = BytesIO()
        response = HttpResponse(content_type="application/pdf")
        response["Content-Disposition"] = f'attachment; filename="Nosozlik_{nosozlik.id}_steps.pdf"'

        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=1*cm, bottomMargin=1*cm)
        styles = getSampleStyleSheet()
        elements = []

        # --- Title ---
        title_text = (
            f"<b><font color='red'>{nosozlik.tarkib.tarkib_raqami if nosozlik.tarkib else '-'}</font></b> harakat tarkibida "
            f"aniqlangan <b><font color='red'>{nosozlik.nosozliklar_haqida.nosozlik_turi if nosozlik.nosozliklar_haqida else '-'}</font></b> "
            f"nosozligi bo‘yicha to‘liq ma’lumot"
        )
        title = Paragraph(
            title_text,
            ParagraphStyle('title', parent=styles['Title'], alignment=1, textColor=colors.darkblue)
        )
        elements.append(title)
        elements.append(Spacer(1, 10))

        # --- Tarkib rasmi ---
        if getattr(nosozlik.tarkib, 'image', None):
            try:
                img = Image(nosozlik.tarkib.image.path, width=15.5*cm, height=6.5*cm)
                img.hAlign = 'CENTER'
                elements.append(img)
                elements.append(Spacer(1, 10))
            except:
                pass

        # --- Ekspluatatsiya jadvali (Texnik Korikdagi kabi) ---
        ekspl_data = [
            [
                Paragraph(f"<font color='red'><b>{h}</b></font>", styles['Normal'])
                for h in ["Ekspluatatsiyaga qo‘yilgan sana", "Turi", "Nomeri", "Masofa", "Hozirgi holati"]
            ],
            [
                Paragraph(f"<b>{nosozlik.tarkib.ishga_tushgan_vaqti.strftime('%d-%m-%Y') if getattr(nosozlik.tarkib, 'ishga_tushgan_vaqti', None) else '-'}</b>", styles['Normal']),
                Paragraph(f"<b>{nosozlik.tarkib.guruhi if getattr(nosozlik.tarkib, 'guruhi', None) else '-'}</b>", styles['Normal']),
                Paragraph(f"<b>{nosozlik.tarkib.tarkib_raqami if getattr(nosozlik.tarkib, 'tarkib_raqami', None) else '-'}</b>", styles['Normal']),
                Paragraph(f"<b>{nosozlik.tarkib.masofa if getattr(nosozlik.tarkib, 'masofa', None) else '-'}</b>", styles['Normal']),
                Paragraph(f"<b>{nosozlik.tarkib.holati if getattr(nosozlik.tarkib, 'holati', None) else '-'}</b>", styles['Normal']),
            ],
        ]
        full_width = A4[0]-2*cm
        col_widths = [full_width*0.18, full_width*0.18, full_width*0.26, full_width*0.18, full_width*0.20]

        ekspl_table = Table(ekspl_data, hAlign='CENTER', colWidths=col_widths)
        ekspl_table.setStyle(TableStyle([
            ('GRID', (0,0), (-1,-1), 0.6, colors.black),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('BACKGROUND', (0,0), (-1,0), colors.whitesmoke),
            ('BOX', (0,0), (-1,-1), 1, colors.black),
            ('FONTSIZE', (0,0), (-1,-1), 9),
        ]))
        elements.append(ekspl_table)
        elements.append(Spacer(1, 12))

        # --- Qizil chiziq ---
        title_line = Table(
            [[""]],
            colWidths=[A4[0] - 2 * cm],
            style=TableStyle([('LINEABOVE', (0,0), (-1,-1), 1, colors.red)])
        )
        elements.append(title_line)
        elements.append(Spacer(1, 6))

        # --- Step-lar ---
        for i, step in enumerate(steps, 1):
            created_by = getattr(step.created_by, 'username', '-')
            step_date = step.created_at.strftime("%d-%m-%Y") if getattr(step, 'created_at', None) else "-"
            kamchilik = getattr(step, "bartaraf_etilgan_nosozliklar", '-')
            ehtiyot_qismlar_detail = ", ".join([
                getattr(eq.ehtiyot_qism, 'ehtiyotqism_nomi', '-')
                for eq in step.ehtiyot_qismlar_step.all()
            ]) or "-"

            step_text = (
                f"<font color='darkblue'>{nosozlik.tarkib.tarkib_raqami}</font> tarkibida aniqlangan "
                f"<font color='red'>{nosozlik.nosozliklar_haqida.nosozlik_turi if nosozlik.nosozliklar_haqida else '-'}</font> "
                f"nosozlik bo‘yicha <font color='red'>{step_date}</font> kuni "
                f"<font color='red'>{created_by}</font> tomonidan qo‘shimcha ma’lumot kiritildi.<br/><br/>"
                f"Depo navbatchisi aniqlagan nosozlik haqida: <font color='red'>{kamchilik}</font><br/>"
                f"Depo navbatchisi ishlatgan ehtiyot qismlar: <font color='red'>{ehtiyot_qismlar_detail}</font>"
            )

            p = Paragraph(
                step_text,
                ParagraphStyle('Normal', alignment=1, leading=12)
            )
            elements.append(p)
            elements.append(Spacer(1, 6))

            step_line = Table(
                [[""]],
                colWidths=[A4[0] - 2 * cm],
                style=TableStyle([('LINEABOVE', (0,0), (-1,-1), 1, colors.red)])
            )
            elements.append(step_line)
            elements.append(Spacer(1, 6))

        # --- Yakuniy akt matni ---
        akt_text = (
            f"<font color='darkblue'>Nosozlik <font color='red'>{nosozlik.nosozliklar_haqida.nosozlik_turi if nosozlik.nosozliklar_haqida else '-'}</font> "
            f"bo‘yicha barcha bosqichlar yakunlandi va AKT ilova qilindi.</font>"
        )
        akt_para = Paragraph(
            akt_text,
            ParagraphStyle('Normal', alignment=1, leading=12)
        )
        elements.append(Spacer(1, 12))
        elements.append(akt_para)
        elements.append(Spacer(1, 6))

        # --- Footer va QR ---
        qr_url = f"https://depo-main.vercel.app/nosozlik/{nosozlik.id}"
        qr = qrcode.QRCode(error_correction=qrcode.constants.ERROR_CORRECT_H, box_size=3, border=1)
        qr.add_data(qr_url)
        qr.make(fit=True)
        qr_buffer = BytesIO()
        qr.make_image(fill_color="black", back_color="white").save(qr_buffer, format='PNG')
        qr_buffer.seek(0)
        qr_img = Image(qr_buffer, width=2.5*cm, height=2.5*cm)
        qr_img.hAlign = 'RIGHT'

        footer_text = Paragraph(
            """<font size=9><b>
            Hujjatning haqiqiyligini tekshirish uchun QR kodni skanerlang.<br/>
            <font color='blue'><b>depo.tm1.uz</b></font> sayti orqali tasdiqlangan.<br/>
            Ushbu hisobotdagi barcha ma’lumotlar uchun xodim mas’uldir.
            </b></font>""",
            ParagraphStyle('footer', parent=styles['Normal'], alignment=0, leading=12),
        )
        footer_table = Table([[footer_text, qr_img]], colWidths=[14*cm, 3*cm])
        footer_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,-1), colors.lightblue),
            ('VALIGN', (0,0), (-1,-1), 'TOP'),
            ('BOX', (0,0), (-1,-1), 0.8, colors.black),
        ]))
        elements.append(footer_table)

        # --- Border ---
        def draw_border(canvas, doc):
            canvas.saveState()
            canvas.setStrokeColor(colors.black)
            canvas.setLineWidth(1)
            canvas.rect(0.8*cm, 0.8*cm, A4[0]-1.6*cm, A4[1]-1.6*cm)
            canvas.restoreState()

        doc.build(elements, onFirstPage=draw_border, onLaterPages=draw_border)
        buffer.seek(0)
        response.write(buffer.getvalue())
        buffer.close()
        return response

    # === EXCEL EXPORT ===
    @action(detail=True, methods=["get"], url_path="export-excel")
    def export_step_excel(self, request, pk=None):
        try:
            nosozlik = Nosozliklar.objects.get(id=pk)
        except Nosozliklar.DoesNotExist:
            return HttpResponse("Bunday nosozlik topilmadi", status=404)

        steps = self.get_queryset().filter(nosozlik_id=pk)
        if not steps.exists():
            return HttpResponse("Bu nosozlik bo‘yicha steplar topilmadi", status=404)

        wb = Workbook()
        ws = wb.active
        ws.title = "Nosozlik Steplar"
        ws.append(["#", "Tarkib raqami", "Nosozlik turi", "Kamchiliklar", "Ehtiyot qismlar", "Xodim", "Sana", "Status"])

        for i, step in enumerate(steps, 1):
            ws.append([
                i,
                nosozlik.tarkib.tarkib_raqami if nosozlik.tarkib else "-",
                nosozlik.nosozliklar_haqida.nosozlik_turi if nosozlik.nosozliklar_haqida else "-",
                step.bartaraf_etilgan_nosozliklar or "-",
                ", ".join([eq.ehtiyot_qism.ehtiyotqism_nomi for eq in step.ehtiyot_qismlar_step.all()]) or "-",
                step.created_by.username if step.created_by else "-",
                step.created_at.strftime("%d-%m-%Y") if step.created_at else "-",
                step.status,
            ])

        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = f'attachment; filename=Nosozlik_{nosozlik.id}_steps.xlsx'
        wb.save(response)
        return response
    
   


class MarshrutJadvalViewSet(viewsets.ModelViewSet):
    queryset = Marshrut.objects.all()
    serializer_class = MarshrutSerializer
    permission_classes = [IsAuthenticated, IsJadvalchi]
    filter_backends = [filters.SearchFilter, DjangoFilterBackend, filters.OrderingFilter]
    pagination_class = CustomPagination
    filterset_fields = ["marshrut_raqam"]
    
    
class YilOyViewSet(viewsets.ModelViewSet):
    queryset = YilOy.objects.all()
    serializer_class = YilOySerializer
    permission_classes = [IsAuthenticated, IsJadvalchi]
    filter_backends = [filters.SearchFilter, DjangoFilterBackend, filters.OrderingFilter]
    pagination_class = CustomPagination
    filterset_fields = ["yil", "oy"]
    
    


class TexnikKorikJadvalViewSet(viewsets.ModelViewSet):
    queryset = TexnikKorikJadval.objects.select_related("tarkib", "tamir_turi", "created_by")
    serializer_class = TexnikKorikJadvalSerializer
    # permission_classes = [IsAuthenticated, IsJadvalchi]
    filter_backends = [filters.SearchFilter, DjangoFilterBackend, filters.OrderingFilter]
    search_fields = ["tarkib__tarkib_raqami", "tamir_turi__tamir_nomi"]
    filterset_fields = ["tarkib__depo", "tamir_turi", "sana", "tarkib"]
    ordering_fields = ["sana", "created_at"]
    pagination_class = CustomPagination
    
    
    def get_permissions(self):
        
        user = self.request.user

        # Agar superuser bo‘lsa, to‘liq ruxsat
        if user.is_superuser:
            permission_classes = [IsAuthenticated]
        # Jadvalchi CRUD qila oladi
        elif getattr(user, "role", None) == "jadval":
            permission_classes = [IsAuthenticated, IsJadvalchi]
        # Texnik va Monitoring faqat o‘qiy oladi
        elif getattr(user, "role", None) in ["texnik", "monitoring"]:
            permission_classes = [IsAuthenticated]
            if self.request.method in SAFE_METHODS:
                # Faqat o‘qish uchun
                permission_classes.append(IsTexnik if user.role == "texnik" else IsMonitoringReadOnly)
            else:
                # POST, PUT, DELETE — ruxsat yo‘q
                self.permission_denied(
                    self.request, message="Sizda bu amalni bajarishga ruxsat yo‘q."
                )
        # Boshqalar hech narsa qila olmaydi
        else:
            self.permission_denied(self.request, message="Sizga ruxsat yo‘q.")

        return [permission() for permission in permission_classes]
    
    def get_queryset(self):
        user = self.request.user
        
        # SuperUser uchun barcha depolardagi yozuvlarni ko'rsatish
        if user.is_superuser:
            queryset = (
                TexnikKorikJadval.objects
                .select_related("tarkib", "tamir_turi", "created_by")
                .filter(tarkib__is_active=True)
            )
        # Oddiy user uchun faqat o'z deposidagi yozuvlarni ko'rsatish
        else:
            user_depo = user.depo
            queryset = (
                TexnikKorikJadval.objects
                .select_related("tarkib", "tamir_turi", "created_by")
                .filter(tarkib__is_active=True, tarkib__depo=user_depo)
            )
        
        # Year filter - ?year=2025
        year = self.request.query_params.get('year')
        if year and year.isdigit():
            queryset = queryset.filter(sana__year=int(year))
            
        # Month filter - ?month=10  
        month = self.request.query_params.get('month')
        if month and month.isdigit():
            queryset = queryset.filter(sana__month=int(month))
            
        # Year-month filter - ?year_month=2025-10
        year_month = self.request.query_params.get('year_month')
        if year_month:
            try:
                year, month = year_month.split('-')
                if year.isdigit() and month.isdigit():
                    queryset = queryset.filter(sana__year=int(year), sana__month=int(month))
            except (ValueError, IndexError):
                pass
            
        return queryset
        

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)
        
    @extend_schema(
        parameters=[
            
            OpenApiParameter(
                name="year_month",
                type=OpenApiTypes.STR,
                location=OpenApiParameter.QUERY,
                description="Yil-oy formatida (masalan: 2025-10)",
                required=False,
            ),
            OpenApiParameter(
                name="tarkib__depo",
                type=OpenApiTypes.INT,
                location=OpenApiParameter.QUERY,
                description="Depo ID sini kiriting",
                required=False,
            ),
            OpenApiParameter(
                name="tamir_turi",
                type=OpenApiTypes.INT,
                location=OpenApiParameter.QUERY,
                description="Tamir turi ID sini kiriting",
                required=False,
            ),
            OpenApiParameter(
                name="sana",
                type=OpenApiTypes.DATE,
                location=OpenApiParameter.QUERY,
                description="Aniq sana (YYYY-MM-DD formatida)",
                required=False,
            ),
        ]
    )
    def list(self, request, *args, **kwargs):
        return super().list(request, *args, **kwargs)
        
        
    @action(detail=False, methods=["get"])
    def export_excel(self, request):
        queryset = self.filter_queryset(self.get_queryset())
        
        data = []
        for j in queryset:
            # None tekshirishlarini qo'shamiz
            tarkib_raqami = j.tarkib.tarkib_raqami if j.tarkib else ""
            depo_nomi = j.tarkib.depo.qisqacha_nomi if j.tarkib and j.tarkib.depo else ""
            tamir_nomi = j.tamir_turi.tamir_nomi if j.tamir_turi else ""
            sana = j.sana.strftime("%d.%m.%Y") if j.sana else ""
            created_by = j.created_by.username if j.created_by else ""
            
            data.append({
                "Tarkib": tarkib_raqami,
                "Depo": depo_nomi,
                "Tamir turi": tamir_nomi,
                "Marshrut": j.marshrut or "",
                "Sana": sana,
                "Yaratgan": created_by,
            })
        
        df = pd.DataFrame(data)
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Texnik Korik')
        
        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="texnik_korik_jadval.xlsx"'
        return response

    @extend_schema(
    parameters=[
        OpenApiParameter(
            name="year",
            type=OpenApiTypes.INT,
            location=OpenApiParameter.QUERY,
            description="Yilni kiriting (masalan: 2025)",
            required=False,
        ),
        OpenApiParameter(
            name="month",
            type=OpenApiTypes.INT,
            location=OpenApiParameter.QUERY,
            description="Oy raqamini kiriting (1-12 oralig‘ida, masalan: 10)",
            required=False,
        ),
    ],
    description="Tanlangan oy va yil uchun texnik ko‘rik jadvalini PDF shaklida eksport qiladi.",
    responses={200: "application/pdf"},
)

    
    @action(detail=False, methods=["get"])
    def export_pdf(self, request):
        yil = int(request.query_params.get("year", datetime.now().year))
        oy = int(request.query_params.get("month", datetime.now().month))
        bugun = datetime.now().strftime("%d.%m.%Y")
        kunlar_soni = calendar.monthrange(yil, oy)[1]

        oy_nomlari = {
            1: "Yanvar", 2: "Fevral", 3: "Mart", 4: "Aprel",
            5: "May", 6: "Iyun", 7: "Iyul", 8: "Avgust",
            9: "Sentabr", 10: "Oktyabr", 11: "Noyabr", 12: "Dekabr"
        }
        oy_nomi = oy_nomlari[oy]

        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=landscape(A3),
            topMargin=1.5 * cm,
            bottomMargin=1.5 * cm,
            leftMargin=1 * cm,
            rightMargin=1 * cm,
        )

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle("Title", parent=styles["Title"], alignment=1, fontSize=20, fontName="Helvetica-Bold")
        depo_title_style = ParagraphStyle("DepoTitle", parent=styles["Normal"], alignment=1, fontSize=14, fontName="Helvetica-Bold")
        boshliq_style = ParagraphStyle("Boshliq", parent=styles["Normal"], alignment=2, fontSize=12, leading=15)
        date_style = ParagraphStyle("Date", parent=styles["Normal"], alignment=2, fontSize=10, leading=14)
        main_text_style = ParagraphStyle("MainText", parent=styles["Normal"], alignment=1, fontSize=13, leading=18)

        # Foydalanuvchi deposiga qarab filter
        user = request.user
        if user.is_superuser:
            barcha_aktiv_tarkiblar = HarakatTarkibi.objects.filter(is_active=True)
        else:
            barcha_aktiv_tarkiblar = HarakatTarkibi.objects.filter(is_active=True, depo=user.depo)

        queryset = self.filter_queryset(self.get_queryset())
        tamirlar = queryset.exclude(tamir_turi__tamir_nomi="TO-1")

        depolar = {}
        for t in barcha_aktiv_tarkiblar:
            depo_nomi = t.depo.qisqacha_nomi if t.depo else "No Depo"
            if depo_nomi not in depolar:
                depolar[depo_nomi] = {}
            depolar[depo_nomi][t.tarkib_raqami] = []

        for j in tamirlar:
            if j.sana.year != yil or j.sana.month != oy:
                continue
            if not j.tarkib or not j.tarkib.depo:
                continue

            depo_nomi = j.tarkib.depo.qisqacha_nomi
            tarkib_raqam = j.tarkib.tarkib_raqami
            tamir_nomi = j.tamir_turi.tamir_nomi if j.tamir_turi else ""

            davom_kun = 1
            if j.tamir_turi and j.tamir_turi.tamirlanish_vaqti == "kun":
                davom_kun = j.tamir_turi.tamirlanish_miqdori or 1
            elif j.tamir_turi and j.tamir_turi.tamirlanish_vaqti == "oy":
                davom_kun = (j.tamir_turi.tamirlanish_miqdori or 1) * 30

            end_kun = min(j.sana.day + davom_kun - 1, kunlar_soni)

            depolar[depo_nomi][tarkib_raqam].append({
                "boshlanish": j.sana.day,
                "tugash": end_kun,
                "tamir_nomi": tamir_nomi,
                "marshrut": j.marshrut or ""
            })

        elements = []

        for depo_nomi, tarkiblar in depolar.items():
            depo_obj = ElektroDepo.objects.filter(qisqacha_nomi=depo_nomi).first()
            depo_full = depo_obj.depo_nomi if depo_obj else depo_nomi
            boshliq_ism = depo_obj.depo_rahbari if depo_obj and depo_obj.depo_rahbari else "__________________"

            boshliq_style = ParagraphStyle(
                'boshliq_style',
                alignment=TA_RIGHT,
                fontSize=12,
                rightIndent=0,
            )

            # TASDIQLAYMAN uchun (sal chaproqda)
            tasdiqlayman_style = ParagraphStyle(
                'tasdiqlayman_style',
                parent=boshliq_style,
                alignment=TA_RIGHT,
                rightIndent=60,  # sal chaproqda
                fontSize=12,
                spaceAfter=3,
            )

            # Depo boshlig‘i qatori uchun (sal o‘ngroqda)
            boshliq_nom_style = ParagraphStyle(
                'boshliq_nom_style',
                parent=boshliq_style,
                alignment=TA_RIGHT,
                rightIndent=30,  
                fontSize=12,
                spaceAfter=4,
            )

            # === 1️ Yuqori o‘ngda — tasdiqlovchi qism ===
            elements.append(Paragraph("<b>TASDIQLAYMAN</b>", tasdiqlayman_style))
            elements.append(Paragraph(f"{depo_full} elektrodeposi boshlig‘i", boshliq_nom_style))
            elements.append(Paragraph(f"_________________{boshliq_ism}", boshliq_style))

            #  Sana formati — ham 10 pt
            hozirgi_yil = datetime.now().year
            sana_matn = f"{hozirgi_yil}-yil &quot;________&quot; _______________"
            elements.append(Paragraph(f"<i>{sana_matn}</i>", ParagraphStyle(
                'sana_style',
                parent=boshliq_style,
                spaceBefore=6  
            )))
            elements.append(Spacer(1, 20))

            # === 2️ Jadval sarlavhasi markazda ===
            elements.append(Paragraph(
                f"<b>{depo_full} ( {depo_nomi} ) elektrodeposidagi harakat tarkiblarining "
                f"{yil} yil {oy_nomi} oyidagi rejali yo‘nalishga chiqish va texnik ko‘riklar o‘tkazish jadvali</b>",
                main_text_style
            ))
            elements.append(Spacer(1, 15))

            # === 3️ Jadval ===
            kunlar = list(range(1, kunlar_soni + 1))
            header = ["Tarkib raqami \\ Sana"] + [str(k) for k in kunlar]
            data = [header]

            table_style = TableStyle([
                ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
                ("BACKGROUND", (1, 0), (-1, 0), colors.lightgreen),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 10),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ])

            row_index = 1
            for tarkib, tamirlar_list in sorted(tarkiblar.items()):
                row = [Paragraph(f"<b>{tarkib}</b>", ParagraphStyle("", fontSize=10, fontName="Helvetica-Bold"))] + [""] * kunlar_soni

                for t in tamirlar_list:
                    start, end, tamir_nomi, marshrut = (
                        t["boshlanish"], t["tugash"], t["tamir_nomi"], t["marshrut"]
                    )
                    if not tamir_nomi and marshrut:
                        row[start] = Paragraph(f"<b>{marshrut}</b>", ParagraphStyle("", fontSize=9, alignment=1, fontName="Helvetica"))
                    elif tamir_nomi:
                        row[start] = Paragraph(f"<b>{tamir_nomi}</b>", ParagraphStyle("", fontSize=8, alignment=1, fontName="Helvetica-Bold"))
                        if tamir_nomi.startswith("TO"):
                            table_style.add("BACKGROUND", (start, row_index), (end, row_index), colors.HexColor("#fff566"))  # sariq
                        elif tamir_nomi.startswith("TR"):
                            table_style.add("BACKGROUND", (start, row_index), (end, row_index), colors.HexColor("#b7ebff"))  # ko‘k
                        else:
                            table_style.add("BACKGROUND", (start, row_index), (end, row_index), colors.HexColor("#ffccc7"))  # qizil

                        table_style.add("SPAN", (start, row_index), (end, row_index))

                data.append(row)
                row_index += 1

            col_widths = [4.0 * cm] + [1.15 * cm] * kunlar_soni
            row_height = 0.9 * cm
            table = Table(data, repeatRows=1, colWidths=col_widths, hAlign="CENTER", rowHeights=row_height)
            table.setStyle(table_style)
            elements.append(table)

            elements.append(Spacer(1, 25))
            if depo_nomi != list(depolar.keys())[-1]:
                elements.append(PageBreak())

        # PDF yaratish
        doc.build(elements)
        buffer.seek(0)
        response = HttpResponse(buffer.getvalue(), content_type="application/pdf")
        response["Content-Disposition"] = f'attachment; filename="texnik_korik_{yil}_{oy}.pdf"'
        return response










